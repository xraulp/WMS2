from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.http import require_POST, require_GET
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.utils import timezone
from django.db.models import Q, Count, Sum, Max, F, OuterRef, Subquery
from django.db import connection, transaction
from io import BytesIO
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
import calendar
from django.http import Http404
from django.utils.text import slugify
import os, json, zipfile, re, logging

from .models import (WarehouseOperation, Catalog, OperationDocument, UserProfile,
                     DeletionLog, DocumentSequence, Tenant, Subscription,
                     NotificationLog, PlatformUser, PLATFORM_ROLE_CHOICES,
                     CATALOG_SCOPES, catalog_scope_of, Invoice,
                     Conversation, ConversationRead, Message,
                     LADO_TENANT, LADO_CLIENTE)
from .utils import generate_pdf_report, generate_label_pdf, generar_pdf_factura
from .almacen import url_firmada
from . import notifications

logger = logging.getLogger(__name__)

# Plazo que propone el formulario de facturacion. Es solo la sugerencia de
# la fecha que aparece escrita: quien emite puede cambiarla.
DIAS_DE_VENCIMIENTO = 15

now_local = timezone.localtime(timezone.now())
generated_at = now_local.strftime('%Y-%m-%d %H:%M')

##### 072526 12:33 Función auxiliar para obtener el tenant
def get_tenant_or_404(request):
    """Devuelve el tenant del request o lanza 404."""
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        raise Http404("Tenant no encontrado")
    return tenant
##### 072526 12:33

# ── PERMISSION HELPERS ────────────────────────────────────────────────────────

def get_profile(user):
    """
    El perfil de quien esta navegando, sin regalarle un rol si no lo tiene.

    Antes fabricaba el perfil que faltaba y le ponia 'superadmin' al
    superusuario de Django y 'manager' a cualquier otro. Lo segundo era lo
    grave: bastaba con existir en `auth_user` y abrir el subdominio de una
    empresa para quedar de manager en ella, con un perfil que ademas quedaba
    escrito en la base como si alguien lo hubiera decidido.

    Ahora quien no tiene perfil recibe uno vacio y sin guardar: todos los
    predicados dan False y las pantallas lo rechazan una por una. Fail-closed,
    y sin filas nuevas. Dar de alta a alguien es un acto explicito de la
    pestana Users.
    """
    try:
        return user.profile
    except UserProfile.DoesNotExist:
        return UserProfile(user=user, role='')

def is_home(user):
    return get_profile(user).is_home()

def is_customer_user(user):
    return get_profile(user).is_customer()

def customer_ops_filter(user, qs):
    profile = get_profile(user)
    if not profile.is_customer():
        return qs
    if profile.customer:
        return qs.filter(
            Q(customer=profile.customer) |
            Q(customer_name_manual__iexact=profile.customer.name)
        )
    return qs.none()


def customer_can_access_op(user, op):
    """
    Version por-objeto de customer_ops_filter, para las vistas que reciben un pk.
    El scoping por tenant ya lo hizo get_object_or_404; esto restringe ademas al
    cliente (Tenant nivel 2) dentro del tenant.

    Los usuarios internos del tenant pasan siempre. Un usuario 'customer' solo
    accede a las operaciones de su propio cliente y, si no tiene cliente asignado,
    no accede a ninguna: mismo criterio fail-closed que customer_ops_filter, para
    que lo que se puede abrir coincida exactamente con lo que se puede listar.
    """
    profile = get_profile(user)
    if not profile.is_customer():
        return True
    if not profile.customer_id:
        return False
    if op.customer_id == profile.customer_id:
        return True
    # customer_ops_filter tambien acepta el nombre capturado a mano; si no lo
    # replicamos aqui, el customer veria la operacion en la lista pero recibiria
    # 403 al abrirla.
    manual = (op.customer_name_manual or '').strip().lower()
    return bool(manual) and manual == profile.customer.name.strip().lower()


# ── AUTH ──────────────────────────────────────────────────────────────────────

def _destino_tras_entrar(user):
    """
    A donde va cada quien al entrar.

    Un usuario de plataforma no pertenece a ninguna empresa, asi que el tablero
    del tenant le daria 404: su sitio es la pantalla de plataforma. Quien tenga
    las dos cosas -hoy, el superusuario, que ademas es admin de su empresa- sigue
    entrando al tablero de siempre y llega a la plataforma por su pestana.
    """
    tiene_empresa = UserProfile.objects.filter(user=user, tenant__isnull=False).exists()
    if not tiene_empresa and platform_role(user):
        return 'platform_dashboard'
    return 'dashboard'


def login_view(request):
    if request.user.is_authenticated:
        return redirect(_destino_tras_entrar(request.user))
    if request.method == 'POST':
        user = authenticate(request,
                            username=request.POST.get('username'),
                            password=request.POST.get('password'))
        if user:
            login(request, user)
            destino = _destino_tras_entrar(user)
            if request.headers.get('HX-Request'):
                url = '/platform/' if destino == 'platform_dashboard' else '/dashboard/'
                return HttpResponse(headers={'HX-Redirect': url})
            return redirect(destino)
        if request.headers.get('HX-Request'):
            return render(request, 'warehouse/partials/login_error.html')
        return render(request, 'warehouse/login.html', {'error': 'Invalid credentials.'})
    return render(request, 'warehouse/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ── DASHBOARD ─────────────────────────────────────────────────────────────────

#@login_required
#def dashboard(request):
    #profile = get_profile(request.user)

#### 072526 12:40 Paso 3: Modificar las vistas de listado (filtro por tenant)
####3.1. dashboard
@login_required
def dashboard(request):
    tenant = get_tenant_or_404(request) #### 072526 12:40 Paso 3: Modificar las vistas de listado (filtro por tenant)
    profile = get_profile(request.user)

    # Filtrar operaciones por tenant
    ops = WarehouseOperation.objects.filter(tenant=tenant).select_related( #### 072526 12:40 Paso 3: Modificar las vistas de listado (filtro por tenant)
        'customer', 'shipper', 'carrier', 'bundle_type', 'created_by'
    ).order_by('-date')

###072526 12:40

    # Obtener lista de usuarios para el filtro (solo para superadmin/home/manager)
    # Acotada a la empresa: el desplegable listaba a todos los usuarios activos
    # del sistema, asi que el nombre de usuario de una empresa aparecia en la
    # pantalla de otra. Es el mismo hueco que tenia `operations_by_user`.
    users = []
    if profile.is_superadmin() or profile.is_home() or profile.is_manager():
        users = User.objects.filter(
            is_active=True, profile__tenant=tenant).order_by('username')
    else:
        users = User.objects.filter(pk=request.user.pk)

    def cat_json(category):
        return json.dumps([
            {'id': e.pk, 'name': e.name}
            for e in Catalog.objects.filter(category=category, active=True, tenant=tenant).order_by('name')
        ])

    ops = anotar_hilos(request.user, customer_ops_filter(request.user, ops)[:200])

    context = {
        'operations': ops,
        # El catalogo va en dos pantallas: la operativa, que mantiene quien
        # captura, y la de clientes, que es del administrador de la empresa.
        'operational_entries': _catalog_entries(tenant, 'operational'),
        'customer_entries':    _catalog_entries(tenant, 'customers'),
        'can_edit_operational': profile.can_edit_catalog('SHIPPER'),
        'can_edit_customers':   profile.can_edit_catalog('CUSTOMER'),
        'customers_json':    cat_json('CUSTOMER'),
        'shippers_json':     cat_json('SHIPPER'),
        'carriers_json':     cat_json('CARRIER'),
        'bundle_types_json': cat_json('BUNDLE_TYPE'),
        'active_tab': request.GET.get('tab', 'form'),
        'profile': profile,
        'is_home': profile.is_home(),
        'users': users,
        # La barra superior llevaba el nombre de una empresa escrito a mano.
        'tenant': tenant,
        # Quien ademas administra la plataforma llega a ella por esta pestana;
        # los usuarios de plataforma que no pertenecen a ninguna empresa ni
        # siquiera pasan por aqui, tienen su propia pantalla.
        'platform_role': platform_role(request.user),
    }
    return render(request, 'warehouse/dashboard.html', context)


# ── OPERATIONS BY USER ─────────────────────────────────────────────────────────

@login_required
@require_GET
def operations_by_user(request, user_id):
    tenant = get_tenant_or_404(request) ###### 3.1. dashboard 072526 20:43
    """Filtrar operaciones creadas por un usuario específico"""
    profile = get_profile(request.user)
    if not profile.is_superadmin() and not profile.is_home() and not profile.is_manager():
        return HttpResponse('Permission denied.', status=403)

    # `profile__tenant=tenant` para no confirmar la existencia de usuarios de
    # otras empresas: el pk venia sin acotar y la vista devolvia su username.
    target_user = get_object_or_404(User, pk=user_id, profile__tenant=tenant)
    # El filtro por usuario estaba comentado, asi que la vista devolvia todas
    # las operaciones del tenant sin importar a quien se le pedia filtrar.
    ops = WarehouseOperation.objects.filter(
        tenant=tenant, created_by=target_user).select_related(
        'customer', 'shipper', 'carrier', 'bundle_type', 'created_by')[:200]

    return render(request, 'warehouse/partials/operations_table.html', {
        'operations': anotar_hilos(request.user, ops),
        'is_home': profile.is_home(),
        'profile': profile,
        'filter_user': target_user.username,
    })


# ── EMAIL HELPERS ─────────────────────────────────────────────────────────────
# El envío vive ahora en warehouse/notifications.py, que además deja constancia
# de cada aviso en NotificationLog. Aquí solo quedan los alias que usan las
# plantillas y las vistas.

_build_subject  = notifications.build_subject
_get_cc_emails  = notifications.get_cc_emails


# ── OPERATION CREATE ──────────────────────────────────────────────────────────

@login_required
@require_POST
def operation_create(request):
    tenant = get_tenant_or_404(request) ############ Usa tenant=tenant al crear la operación 072526 20:49
    profile = get_profile(request.user)
    if not profile.can_create_operations():
        return HttpResponse('<div class="msg-error">✗ Permission denied.</div>', status=422)

    p = request.POST
    op_type  = p.get('operation_type', '').strip()
    date_str = p.get('date', '').strip()

    if op_type not in ('ENTRY', 'EXIT'):
        return HttpResponse('<div class="msg-error">✗ Type of Operation is required.</div>', status=422)
    if not date_str:
        return HttpResponse('<div class="msg-error">✗ Date is required.</div>', status=422)

    from datetime import datetime
    try:
        op_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse('<div class="msg-error">✗ Invalid date format.</div>', status=422)

    def get_catalog(pk_str, category):
        # Filtra por tenant: sin eso, un pk de otra empresa se aceptaba y la
        # operacion quedaba apuntando al catalogo de un tercero — con lo que el
        # aviso de alta se le habria ido a su correo.
        try:
            return Catalog.objects.get(pk=int(pk_str), tenant=tenant,
                                       category=category, active=True)
        except (ValueError, TypeError, Catalog.DoesNotExist):
            return None

    def to_int(val):
        try: return int(val) if val and str(val).strip() else None
        except: return None

    def to_dec(val):
        try:
            from decimal import Decimal
            return Decimal(val) if val and str(val).strip() else None
        except: return None

    customer_obj = get_catalog(p.get('customer_id'), 'CUSTOMER')
    customer_manual = p.get('customer_text', '').strip() if not customer_obj else ''

    required_errors = []
    if not customer_obj and not customer_manual:
        required_errors.append('Customer')
    if not p.get('shipper_id') and not p.get('shipper_text','').strip():
        required_errors.append('Shipper')
    if not p.get('carrier_id') and not p.get('carrier_text','').strip():
        required_errors.append('Carrier')
    if not p.get('bundle_type_id') and not p.get('bundle_type_text','').strip():
        required_errors.append('Bundle Type')
    if not p.get('bundle_qty','').strip():
        required_errors.append('Bundle Qty')
    if not p.get('weight_lbs','').strip() and not p.get('weight_kgs','').strip():
        required_errors.append('Weight (LBS or KGS)')
    if not p.get('description','').strip():
        required_errors.append('Description')

    if required_errors:
        fields = ', '.join(required_errors)
        err_html = (
            '<div class="msg-error" id="op-err" '
            'style="display:flex;align-items:flex-start;justify-content:space-between;gap:10px">'
            '<span>&#10007; Required fields missing: <strong>' + fields + '</strong></span>'
            '<button type="button" onclick="this.parentElement.style.display=\'none\'" '
            'style="background:none;border:none;cursor:pointer;font-size:18px;color:#991b1b;'
            'padding:0 4px;flex-shrink:0;line-height:1">&#x2715;</button>'
            '</div>'
        )
        return HttpResponse(err_html, status=422)

    shipper_obj     = get_catalog(p.get('shipper_id'),     'SHIPPER')
    carrier_obj     = get_catalog(p.get('carrier_id'),     'CARRIER')
    bundle_type_obj = get_catalog(p.get('bundle_type_id'), 'BUNDLE_TYPE')

    op = WarehouseOperation(
        tenant=tenant,  # <--- AGREGAR ESTA LÍNEA 072526 19:51 4.1. operation_create Al final, al crear la operación, asigna el tenant:
        date=op_date, operation_type=op_type,
        entry_dispatched=p.get('entry_dispatched', '').strip(),
        customer=customer_obj, customer_name_manual=customer_manual,
        shipper=shipper_obj,
        shipper_name_manual=p.get('shipper_text','').strip() if not shipper_obj else '',
        invoice=p.get('invoice','').strip(), po_order=p.get('po_order','').strip(),
        seal=p.get('seal','').strip(), carrier=carrier_obj,
        carrier_name_manual=p.get('carrier_text','').strip() if not carrier_obj else '',
        pro=p.get('pro','').strip(), trailer=p.get('trailer','').strip(),
        bundle_type=bundle_type_obj,
        bundle_type_manual=p.get('bundle_type_text','').strip() if not bundle_type_obj else '',
        bundle_qty=to_int(p.get('bundle_qty')),
        weight_lbs=to_dec(p.get('weight_lbs')), weight_kgs=to_dec(p.get('weight_kgs')),
        description=p.get('description','').strip(), note=p.get('note','').strip(),
        damage=bool(p.get('damage')), damage_description=p.get('damage_description','').strip(),
        created_by=request.user,
        # NUEVOS CAMPOS
        ref_aa=p.get('ref_aa', '').strip(),
        ref_dys=p.get('ref_dys', '').strip(),
        pedimento=p.get('pedimento', '').strip(),
    )
    op.save()

    def guess_type(name):
        ext = name.rsplit('.',1)[-1].lower() if '.' in name else ''
        if ext in ('jpg','jpeg','png','gif','webp','heic'): return 'PHOTO'
        if ext in ('mp4','mov','avi','mkv','webm'):         return 'VIDEO'
        if ext in ('pdf','doc','docx','xls','xlsx','csv'): return 'DOCUMENT'
        return 'OTHER'

    # El `tenant` no se pasaba aqui -- si en la subida digital --, asi que todo
    # documento adjuntado al crear la operacion quedaba con el campo en NULL.
    # Es lo que le paso a cinco de los nueve documentos que habia en la base.
    for f in request.FILES.getlist('photos'):
        OperationDocument.objects.create(
            tenant=tenant, operation=op, file_type=guess_type(f.name),
            file=f, original_name=f.name, uploaded_by=request.user)
    for f in request.FILES.getlist('documents'):
        OperationDocument.objects.create(
            tenant=tenant, operation=op, file_type=guess_type(f.name),
            file=f, original_name=f.name, uploaded_by=request.user)

    released_entries = []
    if op.operation_type == 'EXIT' and op.entry_dispatched:
        for eid in [x.strip() for x in op.entry_dispatched.split(',') if x.strip()]:
            try:
                entry_op = WarehouseOperation.objects.get(tenant=tenant, custom_id=eid)
                entry_op.entry_dispatched = op.custom_id
                entry_op.save(update_fields=['entry_dispatched'])
                released_entries.append(entry_op)
            except WarehouseOperation.DoesNotExist:
                pass

    send_wa = p.get('send_whatsapp') == '1'
    email_sent, email_error = notifications.notify_operation_created(
        op, triggered_by=request.user, force_whatsapp=send_wa)

    # Aviso de mercancía liberada, a los clientes de las entradas que esta salida
    # despacha. Se pasa lo que ya se envió arriba para no escribirle dos veces a
    # quien es cliente de la salida y de la entrada a la vez.
    if released_entries:
        already = notifications.email_recipients(notifications.resolve_customer(op)) if email_sent else []
        for entry_op in released_entries:
            notifications.notify_goods_released(
                entry_op, exit_op=op, triggered_by=request.user,
                already_notified=already)

    smtp_not_configured = email_error and any(
        x in str(email_error) for x in ['getaddrinfo','Connection refused','tuservidor'])
    if smtp_not_configured:
        email_sent, email_error = False, 'smtp_not_configured'

    ops = WarehouseOperation.objects.filter(tenant=tenant).select_related(
        'customer','shipper','carrier','bundle_type')
    ops = customer_ops_filter(request.user, ops)[:200]
    return render(request, 'warehouse/partials/operation_success.html', {
        'operation': op, 'operations': ops,
        'email_sent': email_sent, 'email_error': email_error,
        'has_whatsapp': bool(notifications.whatsapp_number(notifications.resolve_customer(op))),
    })


# ── OPERATION DETAIL / DELETE / PDF / LABEL ───────────────────────────────────

@login_required
def operation_detail(request, pk):
    tenant = get_tenant_or_404(request) #### Usa tenant=tenant en get_object_or_404 072526 20:57
    op = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant) ##### 5.1. operation_detail 072526 20:03
    if not customer_can_access_op(request.user, op):
        return HttpResponse('Permission denied.', status=403)

    fields = [
        ('Date',             op.date.strftime('%Y-%m-%d')),
        ('Type',             op.get_operation_type_display()),
        ('Custom ID',        op.custom_id),
        ('Status',           op.status),
        ('Customer',         op.get_customer_display()),
        ('Shipper',          op.get_shipper_display()),
        ('Entries Disp.',    op.entry_dispatched),
        ('Invoice',          op.invoice), ('PO / Order', op.po_order),
        ('Seal',             op.seal),    ('Carrier',    op.get_carrier_display()),
        ('PRO',              op.pro),     ('Trailer',    op.trailer),
        ('Bundle Type',      op.get_bundle_type_display_name()),
        ('Bundle Qty',       op.bundle_qty),
        ('Weight LBS',       op.weight_lbs), ('Weight KGS', op.weight_kgs),
        ('Description',      op.description), ('Note', op.note),
        ('Customer Notes',   op.customer_notes),
        ('Damage',           '⚠ YES' if op.damage else 'No'),
        ('Email Sent',       op.email_sent_at.strftime('%Y-%m-%d') if op.email_sent else 'Not sent'),
        ('Created By',       op.created_by.username if op.created_by else '—'),
        ('REF AA',           op.ref_aa or '—'),
        ('DYS',              op.ref_dys or '—'),
        ('PEDIMENTO',        op.pedimento or '—'),
    ]
    return render(request, 'warehouse/partials/operation_detail.html', {
        'operation': op, 'fields': fields,
        'customer_email': op.get_customer_email() or '',
        'email_subject':  _build_subject(op),
        'is_home':        is_home(request.user),
        'profile':        get_profile(request.user),
        # De que lado del hilo habla quien mira, y si esta operacion tiene con
        # quien conversar. Una operacion cuyo cliente se capturo a mano y no
        # esta en el catalogo no tiene usuarios del otro lado, asi que no se
        # ofrece el hilo en vez de ofrecer uno que nadie leeria.
        'mi_lado':        lado_en_el_hilo(request.user),
        'hay_con_quien':  notifications.resolve_customer(op) is not None,
    })


def _log_deletion(op, user, motivo=''):
    DeletionLog.objects.create(
        deleted_by=user,
        kind='OPERATION',
        custom_id=op.custom_id,
        operation_type=op.operation_type,
        operation_date=op.date,
        customer_name=op.get_customer_display(),
        description=op.description or '',
        reason=motivo,
        tenant=op.tenant,
    )


def _log_document_deletion(doc, user, motivo=''):
    """
    Deja constancia de un archivo que se saca del expediente.

    Se escribe al archivar, no al purgar: lo que le importa al administrador es
    que el archivo dejo de estar, y la purga posterior solo destruye lo que ya
    estaba fuera.
    """
    op = doc.operation
    DeletionLog.objects.create(
        deleted_by=user,
        kind='DOCUMENT',
        custom_id=op.custom_id if op else '',
        operation_date=op.date if op else None,
        customer_name=op.get_customer_display() if op else '',
        document_name=doc.digital_name or doc.original_name or '',
        description=doc.original_name or '',
        reason=motivo,
        tenant=doc.tenant,
    )


def _motivo_de(request):
    """
    El motivo del borrado, obligatorio y de una linea.

    Sin motivo la bitacora dice quien borro pero no por que, que es justo lo
    que se pregunta cuando falta un expediente.
    """
    return request.POST.get('delete_reason', '').strip()


@login_required
def operation_delete_confirm(request, pk):
    tenant = get_tenant_or_404(request)
    profile = get_profile(request.user)
    if not profile.can_delete_operations():
        return HttpResponse('Permission denied.', status=403)
    if request.method != 'POST':
        return HttpResponse(status=405)

    password = request.POST.get('confirm_password', '')
    motivo   = _motivo_de(request)
    op = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant)
    ops_qs = WarehouseOperation.objects.filter(tenant=tenant).select_related(
        'customer', 'shipper', 'carrier', 'bundle_type')
    ops_qs = customer_ops_filter(request.user, ops_qs)[:200]

    def negar(mensaje):
        return render(request, 'warehouse/partials/operations_table.html', {
            'operations': ops_qs, 'delete_error': mensaje,
            'is_home': is_home(request.user), 'profile': profile,
        })

    # La contrasena de borrado es ahora el control que autoriza a borrar, asi
    # que se exige siempre y sin salidas laterales. Antes, quien no la tuviera
    # configurada pasaba con su contrasena de sesion -la que ya esta escrita en
    # el navegador- y un superadmin podia usar la de cualquier manager.
    if not profile.delete_password:
        return negar('No tienes contrasena de borrado configurada. '
                     'Pidesela al administrador de la empresa.')
    if not profile.check_delete_password(password):
        return negar(f'Incorrect password. Record #{pk} was NOT deleted.')
    if not motivo:
        return negar('Escribe el motivo del borrado.')

    _log_deletion(op, request.user, motivo)
    op.delete()
    return render(request, 'warehouse/partials/operations_table.html', {
        'operations': ops_qs, 'delete_success': 'Record deleted successfully.',
        'is_home': is_home(request.user), 'profile': profile,
    })


@login_required
def operation_pdf(request, pk):
    tenant = get_tenant_or_404(request) #### Usa tenant=tenant en get_object_or_404 072526 21:00
    op = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant) ##### 5.3. operation_pd  072526 20:09
    if not customer_can_access_op(request.user, op):
        return HttpResponse('Permission denied.', status=403)
    pdf = generate_pdf_report(op)
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="{op.custom_id}.pdf"'
    return resp


@login_required
def operation_label(request, pk):
    tenant = get_tenant_or_404(request) #### Usa tenant=tenant en get_object_or_404 072526 21:00
    op = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant) ####operation_label 072526 20:10
    if not customer_can_access_op(request.user, op):
        return HttpResponse('Permission denied.', status=403)
    pdf = generate_label_pdf(op)
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="{op.custom_id}_label.pdf"'
    return resp


# Cuantas cifras lleva el consecutivo de cada archivo dentro del ZIP.
#
# Los ceros a la izquierda no son cosmetica: sin ellos el ZIP metia los archivos
# en el orden bueno pero el explorador los mostraba alfabeticos -1, 10, 11,
# 2...-, y quien abre la carpeta ve la secuencia rota. En una entrada se
# fotografia la misma pieza varias veces y de ese orden depende que la
# documentacion aduanal salga bien.
#
# El minimo es tres porque una operacion puede pasar de cien fotos, y con dos
# cifras volveria el mismo desorden a partir de la numero 100. Si alguna llegara
# a pasar de mil, el ancho crece solo.
MINIMO_DE_CIFRAS_EN_EL_ZIP = 3


def _ancho_de_numeracion(cantidad):
    return max(MINIMO_DE_CIFRAS_EN_EL_ZIP, len(str(cantidad)))


@login_required
def operation_download_all(request, pk):
    tenant = get_tenant_or_404(request) #### Usa tenant=tenant en get_object_or_404 072526 21:00
    op = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant) ##### operation_download_all 072526 20:11
    if not customer_can_access_op(request.user, op):
        return HttpResponse('Permission denied.', status=403)
    # El orden va explicito ademas del `ordering` del modelo: es el que decide
    # la numeracion de los archivos dentro del ZIP, y eso no debe depender de
    # que nadie toque el Meta mas adelante.
    #
    # `orden` va primero y no se puede omitir: es la posicion puesta a mano
    # desde el panel, y sin ella el ZIP entregaria las fotos como se subieron
    # aunque alguien las hubiera reordenado -- que es justo para lo que se
    # reordenan. Este orden explicito ya se comio esa equivocacion una vez.
    docs = op.documents.order_by('orden', 'uploaded_at', 'pk')

    if not docs.exists():
        return HttpResponse('No files attached.', status=404)

    # Obtener abreviatura del customer
    customer_abbr = get_customer_abbreviation(op)

    # Construir el prefijo del nombre del archivo
    # Formato: ABREV PO/ORDER INVOICE CUSTOM_ID REF_AA REF_DYS PEDIMENTO
    name_parts = [customer_abbr]

    if op.po_order:
        name_parts.append(op.po_order)
    if op.invoice:
        name_parts.append(op.invoice)
    if op.custom_id:
        name_parts.append(op.custom_id)
    if op.ref_aa:
        name_parts.append(op.ref_aa)
    if op.ref_dys:
        name_parts.append(op.ref_dys)
    if op.pedimento:
        name_parts.append(op.pedimento)

    base_name = ' '.join(name_parts).replace('/', '_')

    # Los archivos se agrupan por tipo -foto, documento, video- y no por
    # extension. Agrupar por extension partia la serie en dos en cuanto una
    # foto llegaba como .jpeg y otra como .jpg: cada grupo empezaba a contar
    # por su cuenta y aparecian dos "foto 1" distintas.
    #
    # Dentro de cada tipo se respeta el orden de subida, que es el orden en que
    # se tomaron: en una misma pieza se fotografia la serie o el lote, luego el
    # peso, luego la tabla nutrimental, y la documentacion aduanal se arma
    # siguiendo esa secuencia. El orden lo garantiza el `ordering` del modelo.
    files_by_type = {}
    for doc in docs:
        files_by_type.setdefault(doc.file_type, []).append(doc)

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for tipo, file_list in files_by_type.items():
            ancho = _ancho_de_numeracion(len(file_list))
            for idx, doc in enumerate(file_list, 1):
                ext = os.path.splitext(doc.original_name or doc.file.name)[1].lower()
                new_filename = f"{base_name} {idx:0{ancho}d}{ext}"
                try:
                    # Se lee del storage en vez de usar `doc.file.path`: `path`
                    # no existe con R2, asi que el ZIP salia vacio en produccion
                    # y el error solo aparecia en la consola. Es el mismo fallo
                    # que dejo los correos sin adjuntos.
                    doc.file.open('rb')
                    try:
                        zf.writestr(new_filename, doc.file.read())
                    finally:
                        doc.file.close()
                except Exception as e:
                    logger.warning('No se pudo agregar al ZIP el documento %s: %s',
                                   doc.pk, e)

    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/zip')
    resp['Content-Disposition'] = f'attachment; filename="{base_name}.zip"'
    return resp


@login_required
@require_GET
def document_file(request, doc_pk):
    """
    Entrega un archivo del expediente, comprobando antes quien lo pide.

    Es la puerta que faltaba. Hasta ahora la pantalla enlazaba al bucket
    directamente: el enlace no llevaba credencial, el bucket estaba publicado
    para que eso funcionara, y en consecuencia cualquiera con la ruta se
    llevaba el archivo sin sesion, sin permiso y sin dejar rastro. Aqui pasa lo
    contrario de cada cosa: hay que estar dentro, hay que poder ver la
    operacion, el enlace que se entrega caduca en minutos y la apertura queda
    en el log.

    El aislamiento es el mismo que el de `operation_detail`, y a proposito: si
    alguien puede abrir la operacion puede abrir sus archivos, y si no, no. El
    tenant sale de la operacion y no del documento porque hay documentos
    antiguos con el campo en NULL —la causa esta arreglada, las filas viejas
    no—, y esos tienen que seguir abriendose.

    Un archivo en la papelera solo lo abre quien puede ver la papelera. Para
    los demas ya no existe, que es lo que significa archivarlo.
    """
    tenant = get_tenant_or_404(request)
    doc = get_object_or_404(
        OperationDocument.todos.select_related('operation'),
        pk=doc_pk, operation__tenant=tenant)

    if not customer_can_access_op(request.user, doc.operation):
        return HttpResponse('Permission denied.', status=403)

    profile = get_profile(request.user)
    if doc.en_papelera and not profile.can_see_deletion_log():
        raise Http404('El documento esta archivado')

    # Hay al menos una fila en produccion con archivo registrado y sin objeto
    # en el bucket. Sin este guard el caso acabaria igual en 404 —abrir el
    # archivo revienta y el except de abajo lo convierte—, pero por un camino
    # que pasa antes por firmar y por el log: mas ruidoso y menos legible.
    if not doc.file or not doc.file.name:
        raise Http404('El documento no tiene archivo')

    descargar = request.GET.get('download') == '1'
    nombre = doc.original_name or os.path.basename(doc.file.name)

    logger.info('Documento %s (%s) abierto por %s en %s%s',
                doc.pk, doc.file.name, request.user.username, tenant.subdomain,
                ' [descarga]' if descargar else '')

    firmada = url_firmada(doc.file, descargar_como=nombre if descargar else None)
    if firmada:
        respuesta = redirect(firmada)
        # La URL firmada caduca y este enlace no, asi que el navegador no debe
        # quedarse con el redirect: al volver a entrar tiene que preguntar otra
        # vez y recibir una firma nueva.
        respuesta['Cache-Control'] = 'private, no-store'
        return respuesta

    # Sin almacen que sepa firmar —el sistema de archivos local, en desarrollo
    # y en las pruebas— el archivo sale por aqui. Es mas lento porque pasa por
    # el servidor, y es lo unico que mantiene la pantalla igual sin R2 detras.
    try:
        contenido = doc.file.open('rb')
    except Exception as e:
        logger.warning('No se pudo abrir el archivo del documento %s: %s', doc.pk, e)
        raise Http404('El archivo no esta disponible')
    return FileResponse(contenido, as_attachment=descargar, filename=nombre)


def get_customer_abbreviation(operation):
    """Obtiene la abreviatura del customer"""
    # Opción A: Desde el modelo Catalog (si agregaste el campo abbreviation)
    if operation.customer and hasattr(operation.customer, 'abbreviation') and operation.customer.abbreviation:
        return operation.customer.abbreviation.upper()

    # Opción B: Desde archivo de configuraciones (opcional)
    try:
        from .customer_abbreviations import CUSTOMER_ABBREVIATIONS
        customer_name = operation.get_customer_display()
        if customer_name in CUSTOMER_ABBREVIATIONS:
            return CUSTOMER_ABBREVIATIONS[customer_name]
    except ImportError:
        pass

    # Fallback: primeras 4 letras del nombre
    customer_name = operation.get_customer_display()
    return customer_name[:4].upper() if customer_name and customer_name != '—' else 'UNKN'


@login_required
@require_POST
def operation_send_email(request, pk):
    tenant = get_tenant_or_404(request) ######### Usa tenant=tenant en get_object_or_404 0072526 21:03
    op        = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant)##### operation_send_email  072526 20:13
    if not customer_can_access_op(request.user, op):
        return HttpResponse('<div class="msg-error">✗ Permission denied.</div>', status=403)
    recipient = request.POST.get('recipient_email','').strip()
    subject   = request.POST.get('subject', _build_subject(op))
    message   = request.POST.get('message','')
    if not recipient:
        return HttpResponse('<div class="msg-error">✗ Recipient email required.</div>')
    sent, error = notifications.send_manual_email(
        op, recipient, subject, message_body=message, triggered_by=request.user)
    if sent:
        return HttpResponse(f'<div class="msg-success">✓ Report sent to {recipient}.</div>')
    return HttpResponse(f'<div class="msg-error">✗ Email failed: {error}</div>')


@login_required
@require_POST
def operation_send_whatsapp(request, pk):
    tenant = get_tenant_or_404(request) ######### Usa tenant=tenant en get_object_or_404 0072526 21:03
    """Send WhatsApp message for a specific operation."""
    op = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant) ##### operation_send_whatsapp 072526 20:13
    if not customer_can_access_op(request.user, op):
        return HttpResponse('Permission denied.', status=403)
    sent, error = notifications.send_manual_whatsapp(op, triggered_by=request.user)
    wa = notifications.whatsapp_number(notifications.resolve_customer(op))
    if sent:
        return HttpResponse(f'<div class="msg-success">✓ WhatsApp sent to {wa}.</div>')
    # Antes esta vista decía "sent" siempre que el cliente tuviera número, aunque
    # el envío hubiera reventado: el error se perdía dentro de _send_whatsapp.
    if error == 'no_number':
        return HttpResponse('<div class="msg-error">✗ No WhatsApp number for this customer.</div>')
    if error == 'twilio_not_configured':
        return HttpResponse('<div class="msg-error">✗ WhatsApp no está configurado en el servidor.</div>')
    return HttpResponse(f'<div class="msg-error">✗ WhatsApp failed: {error}</div>')


# ── SEARCH ────────────────────────────────────────────────────────────────────

@login_required
@require_GET
def operations_search(request):
    tenant = get_tenant_or_404(request) #### 072526 12:51 Modifica la consulta inicial:
    q   = request.GET.get('q','').strip()
    status_filter = request.GET.get('status','').strip()
    #####ops = WarehouseOperation.objects.select_related(
    #####    'customer','shipper','carrier','bundle_type').all()
    #####ops = customer_ops_filter(request.user, ops)
    ops = WarehouseOperation.objects.filter(tenant=tenant).select_related(
        'customer', 'shipper', 'carrier', 'bundle_type' #### 072526 12:54 Modifica la consulta inicial:
    )
    ops = customer_ops_filter(request.user, ops)

    if q:
        ops = ops.filter(
            Q(custom_id__icontains=q) | Q(operation_type__icontains=q) |
            Q(customer__name__icontains=q) | Q(customer_name_manual__icontains=q) |
            Q(shipper__name__icontains=q) | Q(carrier__name__icontains=q) |
            Q(invoice__icontains=q) | Q(po_order__icontains=q) |
            Q(pro__icontains=q) | Q(trailer__icontains=q) |
            Q(description__icontains=q) | Q(date__icontains=q)
        )
    ops_list = list(ops[:500])
    if status_filter in ('Released Goods', 'In Warehouse'):
        ops_list = [o for o in ops_list if o.status == status_filter]
    profile = get_profile(request.user)
    return render(request, 'warehouse/partials/operations_table.html',
                  {'operations': anotar_hilos(request.user, ops_list[:200]),
                   'search_query': q,
                   'is_home': is_home(request.user), 'profile': profile})


# ── FREE ENTRIES ──────────────────────────────────────────────────────────────

@login_required
@require_GET
def free_entries(request):
    tenant = get_tenant_or_404(request) ##### 072526 13:00 3.3. free_entries
    customer_id = request.GET.get('customer_id','').strip()
    if not customer_id:
        return JsonResponse([], safe=False)
    try:
        cid = int(customer_id)
    except (ValueError, TypeError):
        return JsonResponse([], safe=False)
    ops = WarehouseOperation.objects.filter(
        tenant=tenant, ##### 072526 13:00 3.3. free_entries
        operation_type='ENTRY', customer_id=cid,
    ).filter(Q(entry_dispatched__isnull=True)|Q(entry_dispatched='')).order_by('-date')
    return JsonResponse(
        [{'custom_id': op.custom_id, 'date': str(op.date), 'po_order': op.po_order or ''} for op in ops[:100]],
        safe=False)


# ── DIGITAL TAB ───────────────────────────────────────────────────────────────

@login_required
def _expedientes_con_archivos(request, tenant, limite=30):
    """
    Los expedientes que tienen algo guardado, del mas reciente al mas viejo.

    La pantalla solo sabia abrir un expediente si le tecleaban el Custom ID
    completo y exacto, de modo que no habia manera de saber cuales tienen
    archivos: quien no se supiera el numero de memoria se quedaba mirando un
    recuadro vacio. Esto es el indice que faltaba.

    Cuenta solo los documentos vivos; los de la papelera no engordan el numero
    ni asoman un expediente que se quedo sin nada.
    """
    qs = customer_ops_filter(request.user, WarehouseOperation.objects.filter(tenant=tenant))
    return (qs.annotate(
                num_docs=Count('documents',
                               filter=Q(documents__deleted_at__isnull=True)))
              .filter(num_docs__gt=0)
              .order_by('-date', '-id')[:limite])


def digital_search(request):
    tenant = get_tenant_or_404(request) ##### 072526 13:07
    q  = request.GET.get('q','').strip()
    op = None
    candidatos = []
    if q:
        try:
            op = WarehouseOperation.objects.get(tenant=tenant, custom_id__iexact=q) ##### tenant=tenant 072526 13:07
            if not customer_can_access_op(request.user, op):
                op = None
        except WarehouseOperation.DoesNotExist:
            op = None

        if op is None:
            # Exacto o nada era demasiado estricto para un identificador que
            # nadie se aprende: se busca por partes y, si solo hay uno, se abre.
            parciales = customer_ops_filter(
                request.user,
                WarehouseOperation.objects.filter(tenant=tenant, custom_id__icontains=q))
            parciales = (parciales.annotate(
                            num_docs=Count('documents',
                                           filter=Q(documents__deleted_at__isnull=True)))
                         .order_by('-date', '-id')[:30])
            candidatos = list(parciales)
            if len(candidatos) == 1:
                op, candidatos = candidatos[0], []

    profile = get_profile(request.user)
    return render(request, 'warehouse/partials/digital_panel.html',
                  {'operation': op, 'query': q, 'is_home': is_home(request.user),
                   'profile': profile, 'candidatos': candidatos,
                   'con_archivos': [] if (op or q) else _expedientes_con_archivos(request, tenant)})


def _reservar_consecutivos(tenant, day, cantidad):
    """
    Aparta `cantidad` números para los documentos de hoy y devuelve cuáles son.

    El contador vive en `DocumentSequence` y solo sube. La versión anterior lo
    deducía de los documentos que había en ese momento -primero contándolos,
    luego mirando el mayor-, y las dos formas devuelven números ya usados en
    cuanto alguien borra algo: contar los reparte repetidos entre documentos
    vivos, y mirar el mayor libera el último en cuanto se borra, aunque su
    nombre ya haya salido impreso y adjunto en un correo.

    Se reserva de una sola vez para toda la subida y dentro de una transacción,
    con la fila bloqueada donde el motor lo permite. Sin eso, dos operadores
    subiendo a la vez el mismo día leen el mismo valor y se llevan los mismos
    números.
    """
    if cantidad <= 0:
        return []

    with transaction.atomic():
        fila, recien_creada = DocumentSequence.objects.get_or_create(
            tenant=tenant, day=day)

        if recien_creada:
            # Puente con los expedientes anteriores a este contador: se arranca
            # donde se hubieran quedado, para no repetir sus nombres.
            fila.last_value = _mayor_consecutivo_ya_usado(tenant, day)

        # SQLite no sabe bloquear filas y Django protesta si se le pide; en
        # local da igual, porque serializa las escrituras de todas formas.
        if connection.features.has_select_for_update:
            fila = (DocumentSequence.objects.select_for_update()
                    .get(pk=fila.pk))

        primero = fila.last_value + 1
        fila.last_value += cantidad
        fila.save(update_fields=['last_value'])

    return list(range(primero, primero + cantidad))


def _mayor_consecutivo_ya_usado(tenant, day):
    """
    El número más alto que aparece en los `DDMMAA-N` que ya existen ese día.

    Solo se usa para sembrar el contador la primera vez. Los nombres con otra
    forma se ignoran: en su día se pudieron cargar a mano.

    Cuenta tambien los documentos que estan en la papelera -de ahi `todos`-:
    un nombre que ya salio impreso o adjunto en un correo no puede volver a
    asignarse porque el archivo se haya archivado, y menos aun cuando ese
    archivo puede restaurarse.
    """
    nombres = OperationDocument.todos.filter(
        tenant=tenant, digital_name__startswith=f'{day}-'
    ).values_list('digital_name', flat=True)

    mayor = 0
    for nombre in nombres:
        sufijo = (nombre or '').split('-', 1)[-1]
        if sufijo.isdigit():
            mayor = max(mayor, int(sufijo))
    return mayor


@login_required
@require_POST
def digital_upload(request, pk):
    tenant = get_tenant_or_404(request) ######### Usa tenant=tenant al crear el documento 072526 20:51
    op = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant)
    if not customer_can_access_op(request.user, op):
        return HttpResponse('Permission denied.', status=403)

    from datetime import date as date_cls
    today_str = date_cls.today().strftime('%d%m%y')
    archivos = request.FILES.getlist('files')
    consecutivos = _reservar_consecutivos(tenant, today_str, len(archivos))

    uploaded = []
    # Las posiciones arrancan donde acaba lo que ya hay: un archivo nuevo va al
    # final del expediente, no al principio.
    #
    # Que cada uno de la misma tanda avance una posicion no llega a notarse:
    # con todos en la misma, el desempate por fecha de subida los deja en el
    # mismo sitio. Se hace igual porque dos archivos en la posicion 5 es una
    # mentira que alguien acabara leyendo, y porque asi el orden no depende de
    # una segunda regla.
    posicion = _siguiente_orden(op)
    for f, consecutive in zip(archivos, consecutivos):
        digital_name = f'{today_str}-{consecutive}'
        ext = f.name.rsplit('.',1)[-1].lower() if '.' in f.name else ''
        if ext in ('jpg','jpeg','png','gif','webp','heic'):  ftype = 'PHOTO'
        elif ext in ('mp4','mov','avi','mkv','webm'):         ftype = 'VIDEO'
        elif ext in ('pdf','doc','docx','xls','xlsx','csv'): ftype = 'DOCUMENT'
        else:                                                  ftype = 'OTHER'

        doc = OperationDocument.objects.create(
            tenant=tenant,  # <--- AGREGAR ESTA LÍNEA 4.2. digital_upload Al crear el documento, asignar tenant: 072526 19:55
            operation=op, file_type=ftype, file=f,
            original_name=f.name, digital_name=digital_name,
            uploaded_by=request.user, orden=posicion)
        posicion += 1
        uploaded.append(doc)

    profile = get_profile(request.user)
    # No se le avisa al cliente de los archivos que él mismo acaba de subir.
    if uploaded and not profile.is_customer():
        notifications.notify_documents_added(op, uploaded, triggered_by=request.user)
    return render(request, 'warehouse/partials/digital_panel.html', {
        'operation': op, 'query': op.custom_id,
        'is_home': is_home(request.user),
        'profile': profile,
        'upload_success': f'{len(uploaded)} file(s) uploaded.',
    })

def _siguiente_orden(op):
    """
    La posicion que le toca al proximo archivo de este expediente.

    Hace falta porque un expediente ya reordenado tiene posiciones 1..N, y un
    archivo nuevo con la posicion en cero se colaria **al principio** de la
    secuencia en vez de al final. Lo que se sube despues va despues.
    """
    mayor = op.documents.aggregate(m=Max('orden'))['m'] or 0
    return mayor + 1


def _renumerar_expediente(docs):
    """
    Deja las posiciones en 1..N siguiendo el orden de la lista que se le pasa.

    Renumerar entero -y no solo los dos que se intercambian- es lo que arregla
    de una vez los expedientes viejos, donde todos los documentos valen cero y
    el orden lo lleva la fecha de subida. Solo escribe los que cambian.
    """
    for posicion, doc in enumerate(docs, 1):
        if doc.orden != posicion:
            doc.orden = posicion
            doc.save(update_fields=['orden'])


@login_required
@require_POST
def digital_reorder(request, doc_pk):
    """
    Sube o baja un archivo una posicion dentro de su expediente.

    El orden de las fotos es informacion: en una entrada se fotografia la misma
    pieza varias veces -la serie o el lote, el peso, la tabla nutrimental- y la
    documentacion aduanal se arma siguiendo esa secuencia. El orden de subida
    la conserva cuando las fotos se toman y se suben una a una desde el movil,
    pero no cuando el operador las selecciona de un tiron desde la PC: ahi el
    navegador las manda como le parece y la secuencia nace mal. Hasta ahora la
    unica salida era borrarlas y volver a subirlas.

    Mueve de uno en uno a proposito. Arrastrar seria mas comodo con muchas
    fotos, pero necesita JavaScript que hay que probar en el movil, que es
    donde mas se usa esta pantalla; esto funciona en los dos sitios y no puede
    dejar el expediente a medias.

    Quien puede subir archivos puede ordenarlos: mismo criterio que
    `digital_upload`, para que no haya que explicar dos reglas distintas sobre
    la misma pantalla.
    """
    tenant = get_tenant_or_404(request)
    doc = get_object_or_404(OperationDocument.objects.select_related('operation'),
                            pk=doc_pk, operation__tenant=tenant)
    op = doc.operation
    if not customer_can_access_op(request.user, op):
        return HttpResponse('Permission denied.', status=403)

    direccion = request.POST.get('direccion')
    documentos = list(op.documents.all())
    posiciones = [d.pk for d in documentos]
    # Segunda red: el manager de arriba ya deja fuera lo archivado, y esto lo
    # vuelve a cerrar desde otro lado. Se comprobo cambiando el manager por el
    # que si ve la papelera: el 404 sigue saliendo, por aqui.
    if doc.pk not in posiciones:
        raise Http404('El documento no esta en el expediente')

    actual = posiciones.index(doc.pk)
    destino = actual - 1 if direccion == 'arriba' else actual + 1

    if 0 <= destino < len(documentos):
        documentos[actual], documentos[destino] = documentos[destino], documentos[actual]
        _renumerar_expediente(documentos)
    # Fuera de rango no es un error: es el primero al que le dan a subir. La
    # pantalla se devuelve igual, sin tocar nada.

    op.refresh_from_db()
    return _panel_digital(request, op, get_profile(request.user))


def _delete_stored_file(doc):
    """
    Borra el archivo del storage, sea local o R2.

    La version anterior hacia `os.path.exists(doc.file.path)`, y `path` **no
    existe cuando los archivos viven en R2**: lanza NotImplementedError. En
    `digital_delete_file` un `except: pass` se lo tragaba y el objeto quedaba
    huerfano en el bucket — accesible todavia por su URL publica, porque el
    dominio de R2 sirve los archivos sin pedir credenciales. En
    `digital_delete_multiple` era peor: la excepcion saltaba antes del
    `doc.delete()`, asi que el borrado multiple no borraba nada y reportaba
    error. `FieldFile.delete()` habla con el storage que este configurado.
    """
    if not doc.file:
        return True, None
    try:
        doc.file.delete(save=False)
        return True, None
    except Exception as e:
        logger.warning('No se pudo borrar del storage el archivo del documento %s: %s',
                       doc.pk, e)
        return False, str(e)


def _panel_digital(request, op, profile, status=200, **extra):
    contexto = {
        'operation': op,
        'query': op.custom_id if op else '',
        'is_home': profile.is_home(),
        'profile': profile,
    }
    contexto.update(extra)
    return render(request, 'warehouse/partials/digital_panel.html', contexto,
                  status=status)


def _rechazo_de_borrado(request, op, profile, motivo_del_rechazo):
    """
    Un borrado que no se hizo, con su porque y con un 400.

    El estado importa: la pantalla pintaba el panel devuelto y avisaba "archivo
    enviado a la papelera" pasara lo que pasara, porque un rechazo llegaba con
    el mismo 200 que un exito. Quien tecleaba mal la contrasena -o no la tenia
    configurada, que era el caso mas frecuente- se quedaba con el aviso de que
    todo habia ido bien y el archivo intacto en el expediente.
    """
    return _panel_digital(request, op, profile, status=400,
                          upload_error=motivo_del_rechazo)


@login_required
@require_POST
def digital_delete_file(request, doc_pk):
    """
    Saca un archivo del expediente y lo manda a la papelera.

    No destruye nada: el objeto sigue en R2 y el registro sigue en la base con
    la marca de quien lo quito y por que. Un archivo del expediente ya salio
    impreso y adjunto en un correo, de modo que quitarlo de la vista y
    destruirlo son dos decisiones distintas y de dos personas distintas.
    """
    tenant = get_tenant_or_404(request)
    profile = get_profile(request.user)
    if not profile.can_delete_documents():
        return HttpResponse('Permission denied.', status=403)

    doc = get_object_or_404(OperationDocument.todos, pk=doc_pk, tenant=tenant)
    op = doc.operation
    password = request.POST.get('confirm_password', '')
    motivo   = _motivo_de(request)

    if not profile.delete_password:
        return _rechazo_de_borrado(
            request, op, profile,
            '❌ No tienes contrasena de borrado configurada. Pidesela a un '
            'administrador: se asigna en la pestana Users.')
    if not profile.check_delete_password(password):
        return _rechazo_de_borrado(
            request, op, profile,
            '❌ Contrasena de eliminacion incorrecta. No se pudo eliminar el archivo.')
    if not motivo:
        return _rechazo_de_borrado(
            request, op, profile, '❌ Escribe el motivo del borrado.')

    if not doc.en_papelera:
        doc.archivar(request.user, motivo)
        _log_document_deletion(doc, request.user, motivo)

    return _panel_digital(request, op, profile,
                          upload_success='✓ Archivo enviado a la papelera.')


@login_required
@require_POST
def digital_delete_multiple(request):
    """
    Lo mismo para varios archivos de una vez.
    """
    tenant = get_tenant_or_404(request)
    profile = get_profile(request.user)
    if not profile.can_delete_documents():
        return HttpResponse('Permission denied.', status=403)

    ids_str = request.POST.get('ids', '').strip()
    if not ids_str:
        return HttpResponse('<div class="msg-error">✗ No se seleccionaron archivos.</div>')

    doc_ids = [int(x.strip()) for x in ids_str.split(',') if x.strip().isdigit()]
    if not doc_ids:
        return HttpResponse('<div class="msg-error">✗ IDs inválidos.</div>')

    docs = list(OperationDocument.todos.filter(pk__in=doc_ids, tenant=tenant))
    if not docs:
        return HttpResponse('<div class="msg-error">✗ Ninguno de los archivos seleccionados existe.</div>')

    op = docs[0].operation
    password = request.POST.get('confirm_password', '')
    motivo   = _motivo_de(request)

    if not profile.delete_password:
        return _rechazo_de_borrado(
            request, op, profile,
            '❌ No tienes contrasena de borrado configurada. Pidesela a un '
            'administrador: se asigna en la pestana Users.')
    if not profile.check_delete_password(password):
        return _rechazo_de_borrado(
            request, op, profile,
            '❌ Contrasena de eliminacion incorrecta.')
    if not motivo:
        return _rechazo_de_borrado(
            request, op, profile, '❌ Escribe el motivo del borrado.')

    archivados = 0
    for doc in docs:
        if doc.en_papelera:
            continue
        doc.archivar(request.user, motivo)
        _log_document_deletion(doc, request.user, motivo)
        archivados += 1

    return _panel_digital(
        request, op, profile,
        upload_success='✓ %d archivo(s) enviado(s) a la papelera.' % archivados)


# ── PAPELERA Y BITACORA DE BORRADOS ───────────────────────────────────────────
# Lo que sustituye al permiso denegado es el rastro: quien borra deja constancia
# de que se llevo y por que, y lo que se lleva de un expediente se puede
# devolver. Las dos pantallas viven en la misma pestana.

def _contexto_papelera(request, tenant, profile, **extra):
    archivados = (OperationDocument.todos
                  .filter(tenant=tenant, deleted_at__isnull=False)
                  .select_related('operation', 'deleted_by')
                  .order_by('-deleted_at')[:200])
    registros = (DeletionLog.objects
                 .filter(tenant=tenant)
                 .select_related('deleted_by')[:200])
    contexto = {
        'archivados': archivados,
        'deletion_logs': registros,
        'profile': profile,
        'tenant': tenant,
    }
    contexto.update(extra)
    return contexto


@login_required
def deletion_log(request):
    """
    La bitacora de borrados y la papelera, juntas.

    Hasta ahora `DeletionLog` solo se leia desde el admin de Django siendo
    superusuario, que es justo el acceso que el proyecto esta retirando.
    """
    tenant = get_tenant_or_404(request)
    profile = get_profile(request.user)
    if not profile.can_see_deletion_log():
        return HttpResponse('Permission denied.', status=403)
    return render(request, 'warehouse/partials/deletion_log.html',
                  _contexto_papelera(request, tenant, profile))


@login_required
@require_POST
def document_restore(request, doc_pk):
    tenant = get_tenant_or_404(request)
    profile = get_profile(request.user)
    if not profile.can_see_deletion_log():
        return HttpResponse('Permission denied.', status=403)

    doc = get_object_or_404(OperationDocument.todos, pk=doc_pk, tenant=tenant)
    doc.restaurar()
    return render(request, 'warehouse/partials/deletion_log.html',
                  _contexto_papelera(request, tenant, profile,
                                     aviso='✓ %s volvio al expediente.'
                                           % (doc.digital_name or doc.original_name)))


@login_required
@require_POST
def document_purge(request, doc_pk):
    """
    Destruye de verdad un archivo que ya estaba en la papelera.

    Es lo unico irreversible de esta pantalla, asi que se queda en el
    administrador de la empresa y exige su contrasena de borrado.
    """
    tenant = get_tenant_or_404(request)
    profile = get_profile(request.user)
    if not profile.can_purge_documents():
        return HttpResponse('Permission denied.', status=403)

    doc = get_object_or_404(OperationDocument.todos, pk=doc_pk, tenant=tenant)
    if not doc.en_papelera:
        return render(request, 'warehouse/partials/deletion_log.html',
                      _contexto_papelera(request, tenant, profile,
                                         aviso='⚠ Ese archivo sigue en el expediente.'))

    if not profile.check_delete_password(request.POST.get('confirm_password', '')):
        return render(request, 'warehouse/partials/deletion_log.html',
                      _contexto_papelera(request, tenant, profile,
                                         aviso='❌ Contrasena de eliminacion incorrecta.'))

    nombre = doc.digital_name or doc.original_name
    _delete_stored_file(doc)
    doc.delete()
    return render(request, 'warehouse/partials/deletion_log.html',
                  _contexto_papelera(request, tenant, profile,
                                     aviso='✓ %s se destruyo definitivamente.' % nombre))


# ── REPORT GENERATOR ──────────────────────────────────────────────────────────

@login_required
def report_generator(request):
    tenant    = get_tenant_or_404(request)
    customers = Catalog.objects.filter(category='CUSTOMER', active=True, tenant=tenant).order_by('name')
    profile   = get_profile(request.user)

    users = []
    if profile.is_superadmin() or profile.is_home():
        users = User.objects.filter(is_active=True, profile__tenant=tenant).order_by('username')
    elif profile.is_manager() or profile.is_staff_role():
        users = User.objects.filter(pk=request.user.pk)
    else:
        users = User.objects.filter(pk=request.user.pk)

    if profile.is_customer() and profile.customer:
        customers = customers.filter(pk=profile.customer.pk)

    results  = None
    filters  = {}
    error    = None

    if request.GET.get('search'):
        all_customers  = request.GET.get('all_customers', '')
        customer_ids   = request.GET.getlist('customer_ids')
        created_by_id  = request.GET.get('created_by', '').strip()

        if not customer_ids and not all_customers:
            error = 'Please select at least one customer or choose All Customers.'
        else:
            ####ops = WarehouseOperation.objects.select_related(
            ops = WarehouseOperation.objects.filter(tenant=tenant).select_related(   ######072526 13:13  En report_generator, la consulta base debe ser:
                'customer', 'shipper', 'carrier', 'bundle_type', 'created_by').all()
            ops = customer_ops_filter(request.user, ops)

            date_from    = request.GET.get('date_from', '').strip()
            date_to      = request.GET.get('date_to', '').strip()
            op_type      = request.GET.get('op_type', '').strip()
            undispatched = request.GET.get('undispatched', '')
            status_f     = request.GET.get('status_filter', '').strip()

            if created_by_id and created_by_id.isdigit():
                ops = ops.filter(created_by_id=int(created_by_id))
                filters['created_by_id'] = created_by_id
                try:
                    creator = User.objects.get(pk=int(created_by_id), profile__tenant=tenant)
                    filters['created_by_name'] = creator.username
                except User.DoesNotExist:
                    pass

            if all_customers:
                filters['all_customers'] = True
                filters['customer_label'] = 'All Customers'
                filters['customer_ids_list'] = []
            elif customer_ids:
                cid_ints = [int(x) for x in customer_ids if x.isdigit()]
                ops = ops.filter(customer_id__in=cid_ints)
                filters['customer_ids_list'] = customer_ids
                names = list(Catalog.objects.filter(pk__in=cid_ints, tenant=tenant).values_list('name', flat=True))
                filters['customer_label'] = ', '.join(names)
                if cid_ints:
                    filters['customer_id'] = str(cid_ints[0])

            if date_from:
                ops = ops.filter(date__gte=date_from); filters['date_from'] = date_from
            if date_to:
                ops = ops.filter(date__lte=date_to);   filters['date_to']   = date_to
            if op_type in ('ENTRY', 'EXIT'):
                ops = ops.filter(operation_type=op_type); filters['op_type'] = op_type
            if undispatched == '1':
                ops = ops.filter(operation_type='ENTRY').filter(
                    Q(entry_dispatched__isnull=True) | Q(entry_dispatched=''))
                filters['undispatched'] = True
            if status_f:
                filters['status_filter'] = status_f

            results_list = list(ops.order_by('-date')[:500])
            if status_f in ('Released Goods', 'In Warehouse'):
                results_list = [o for o in results_list if o.status == status_f]
            results = results_list

    return render(request, 'warehouse/partials/report_generator.html', {
        'customers': customers,
        'results': results,
        'filters': filters,
        'is_home': is_home(request.user),
        'error': error,
        'profile': profile,
        'users': users,
    })


@login_required
def report_generator_pdf(request):
    from .utils import generate_operations_report_pdf
    tenant = get_tenant_or_404(request)
    ops_ids = request.GET.get('ids','').split(',')
    ops = WarehouseOperation.objects.filter(pk__in=[i for i in ops_ids if i], tenant=tenant).select_related(
        'customer','shipper','carrier','bundle_type', 'created_by')
    ops = customer_ops_filter(request.user, ops)
    title   = request.GET.get('title', 'Operations Report')
    pdf     = generate_operations_report_pdf(list(ops), title)
    resp    = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="report.pdf"'
    return resp


@login_required
@require_POST
def report_generator_email(request):
    from .utils import generate_operations_report_pdf
    tenant           = get_tenant_or_404(request)
    ids_raw          = request.POST.get('ids', '').strip().rstrip(',')
    title            = request.POST.get('title', 'Operations Report')
    extra_emails_str = request.POST.get('extra_emails', '').strip()
    customer_id      = request.POST.get('customer_id', '').strip()
    all_customers    = request.POST.get('all_customers', '') == '1'

    pk_list = []
    for i in ids_raw.split(','):
        i = i.strip()
        if i.isdigit():
            pk_list.append(int(i))

    if not pk_list:
        return HttpResponse('<div class="msg-error">✗ No records selected.</div>')

    ops = WarehouseOperation.objects.filter(pk__in=pk_list, tenant=tenant).select_related(
        'customer', 'shipper', 'carrier', 'bundle_type', 'created_by').order_by('-date')
    ops = customer_ops_filter(request.user, ops)
    ops_list = list(ops)

    if not ops_list:
        return HttpResponse('<div class="msg-error">✗ No records found.</div>')

    recipients = []

    if not all_customers and customer_id and customer_id.isdigit():
        try:
            cat = Catalog.objects.get(pk=int(customer_id), tenant=tenant)
            if cat.contact_email:
                for addr in cat.contact_email.split(','):
                    addr = addr.strip()
                    if addr and addr not in recipients:
                        recipients.append(addr)
        except (Catalog.DoesNotExist, ValueError):
            pass

    if extra_emails_str:
        for em in extra_emails_str.split(','):
            em = em.strip()
            if em and em not in recipients:
                recipients.append(em)

    if (all_customers or (customer_id and ',' in customer_id)) and not recipients:
        return HttpResponse('<div class="msg-error">✗ Para múltiples clientes, debes ingresar al menos un email en el campo "Email address".</div>')

    if not recipients:
        return HttpResponse('<div class="msg-error">✗ No hay destinatarios. Ingresa un email en el campo correspondiente.</div>')

    try:
        pdf = generate_operations_report_pdf(ops_list, title)
        email = EmailMessage(
            subject=title,
            body=f'Adjunto encontrará el reporte de operaciones solicitado.\n\n'
                 f'Registros incluidos: {len(ops_list)}\n'
                 f'Fecha y hora de generación: {generated_at} (Hora Central)\n\n'
                 f'Saludos cordiales,\n'
                 f'{tenant.name}',
            to=recipients,
            cc=_get_cc_emails(tenant),
        )
        email.attach('report.pdf', pdf, 'application/pdf')
        email.send()
        return HttpResponse(
            f'<div class="msg-success">✓ Reporte con {len(ops_list)} registro(s) enviado a {", ".join(recipients)}.</div>'
        )
    except Exception as e:
        return HttpResponse(f'<div class="msg-error">✗ Error al enviar: {e}</div>')


# ── CATALOG ───────────────────────────────────────────────────────────────────

def _catalog_scope(request, default='operational'):
    """
    Cual de las dos pantallas del catalogo pidio esto.

    Va en el formulario porque no siempre se puede deducir: la importacion por
    Excel trae categorias mezcladas y la vista tiene que saber que tabla
    refrescar. Un valor desconocido cae en la operativa, que es la de menos
    permiso.
    """
    scope = (request.POST.get('scope') or request.GET.get('scope') or '').strip()
    return scope if scope in CATALOG_SCOPES else default


def _catalog_entries(tenant, scope):
    return Catalog.objects.filter(
        active=True, tenant=tenant, category__in=CATALOG_SCOPES[scope]
    ).order_by('category', 'name')


def _catalog_table_id(scope):
    """El id del contenedor que HTMX reemplaza. Uno por pantalla, porque las
    dos tablas conviven en la misma pagina."""
    return 'customer-table' if scope == 'customers' else 'catalog-table'


def _catalog_table_context(request, tenant, scope, **extra):
    profile = get_profile(request.user)
    contexto = {
        'catalog_entries': _catalog_entries(tenant, scope),
        'scope': scope,
        'table_id': _catalog_table_id(scope),
        # Manager y staff ven la pantalla de clientes para consultarla; los
        # botones de alta, edicion y baja solo se pintan a quien puede usarlos.
        'can_edit': profile.can_edit_catalog(CATALOG_SCOPES[scope][0]),
    }
    contexto.update(extra)
    return contexto


@login_required
@require_POST
def catalog_create(request):
    tenant = get_tenant_or_404(request) ###### Usa tenant=tenant al crear el catálogo 072526 20:52
    profile = get_profile(request.user)
    if profile.is_customer():
        return HttpResponse('Permission denied.', status=403)
    p = request.POST
    category = p.get('category','').strip()
    name     = p.get('name','').strip()
    if not category or not name:
        return HttpResponse('<div class="msg-error">✗ Category and Name are required.</div>')
    # La categoría llega de un desplegable del formulario, así que quitarle la
    # opción a quien no debe verla no basta: el POST se manda igual. Dar de alta
    # un cliente queda reservado al administrador de la empresa.
    if not profile.can_edit_catalog(category):
        return HttpResponse(
            '<div class="msg-error">✗ Only an administrator can create customers.</div>',
            status=403)
    entry = Catalog.objects.create(
        tenant=tenant,  # <--- AGREGAR ESTA LÍNEA 4.3. catalog_create 072526 19:57
        category=category, name=name,
        # Los dos formularios que llegan aquí -el del escritorio y el del móvil-
        # pintan la abreviatura y el operador la escribe, pero el alta no la
        # guardaba: se perdía sin avisar y solo se podía poner volviendo a
        # editar el registro. Se guarda None en vez de cadena vacía para que
        # coincida con lo que hace el alta de cliente con usuario.
        abbreviation=p.get('abbreviation','').strip().upper() or None,
        contact_email=p.get('contact_email','').strip(),
        phone=p.get('phone','').strip(),
        address=p.get('address','').strip(),
        notes=p.get('notes','').strip(),
        whatsapp=p.get('whatsapp','').strip(),
    )
    # La tabla que se refresca es la de la pantalla a la que pertenece lo que
    # se acaba de crear, no siempre la misma.
    scope = catalog_scope_of(category)
    table_html = render_to_string(
        'warehouse/partials/catalog_table.html',
        _catalog_table_context(request, tenant, scope), request=request)
    return HttpResponse(
        f'<div class="msg-success">✓ {entry.name} ({entry.get_category_display()}) saved.</div>'
        f'<div id="{_catalog_table_id(scope)}" hx-swap-oob="innerHTML">{table_html}</div>'
    )


@login_required
def catalog_edit(request, pk):
    tenant = get_tenant_or_404(request) #### Usa tenant=tenant en get_object_or_404 072526 20:55
    profile = get_profile(request.user)
    if profile.is_customer():
        return HttpResponse('Permission denied.', status=403)
    entry = get_object_or_404(Catalog, pk=pk, tenant=tenant) #### 4.4. catalog_edit Verificar que el catálogo pertenece al tenant:
    # La categoría no se puede cambiar desde aquí, así que la que manda es la
    # que ya tiene la entrada.
    if not profile.can_edit_catalog(entry.category):
        return HttpResponse('Permission denied.', status=403)
    if request.method == 'POST':
        p = request.POST
        entry.name          = p.get('name', entry.name).strip()
        entry.contact_email = p.get('contact_email', '').strip()
        entry.phone         = p.get('phone', '').strip()
        entry.address       = p.get('address', '').strip()
        entry.notes         = p.get('notes', '').strip()
        entry.whatsapp      = p.get('whatsapp', '').strip()
        entry.abbreviation  = p.get('abbreviation', '').strip().upper() or None
        # Las preferencias de aviso solo existen para los clientes; para el resto
        # de categorías el formulario ni siquiera pinta los checkboxes, así que
        # leerlos ahí apagaría valores que nadie quiso cambiar.
        if entry.category == 'CUSTOMER':
            entry.notify_email        = p.get('notify_email') == 'on'
            entry.notify_whatsapp     = p.get('notify_whatsapp') == 'on'
            entry.notify_on_create    = p.get('notify_on_create') == 'on'
            entry.notify_on_release   = p.get('notify_on_release') == 'on'
            entry.notify_on_documents = p.get('notify_on_documents') == 'on'
        entry.save()
        return render(request, 'warehouse/partials/catalog_table.html',
                      _catalog_table_context(request, tenant,
                                             catalog_scope_of(entry.category),
                                             edit_success=f'{entry.name} updated.'))
    return render(request, 'warehouse/partials/catalog_edit_form.html', {'entry': entry})


@login_required
def catalog_delete(request, pk):
    tenant = get_tenant_or_404(request) ###### Usa tenant=tenant en get_object_or_404 072526 20:56
    profile = get_profile(request.user)
    if profile.is_customer():
        return HttpResponse('Permission denied.', status=403)
    entry = get_object_or_404(Catalog, pk=pk, tenant=tenant) #####4.5. catalog_delete 072526 20:00
    if not profile.can_edit_catalog(entry.category):
        return HttpResponse('Permission denied.', status=403)
    if request.method == 'POST':
        entry.active = False
        entry.save()
        return render(request, 'warehouse/partials/catalog_table.html',
                      _catalog_table_context(request, tenant,
                                             catalog_scope_of(entry.category)))
    return HttpResponse(status=405)


@login_required
def catalog_list(request):
    tenant = get_tenant_or_404(request) ##### 072526 13:20 3.6. catalog_list, catalog_autocomplete
    scope = _catalog_scope(request)
    return render(request, 'warehouse/partials/catalog_table.html',
                  _catalog_table_context(request, tenant, scope))
    ##})


@login_required
def catalog_autocomplete(request):
    tenant = get_tenant_or_404(request)  ##### 072526 13:17 3.6. catalog_list, catalog_autocomplete
    category = request.GET.get('category','')
    q        = request.GET.get('q','')
    entries  = Catalog.objects.filter(tenant=tenant, category=category, active=True) ##### 072526 tenant=tenant,
    if q: entries = entries.filter(name__icontains=q)
    return JsonResponse([{'id': e.pk, 'name': e.name} for e in entries[:20]], safe=False)


# ── USER MANAGEMENT ───────────────────────────────────────────────────────────

def _perfil_de(user):
    """
    El perfil de otro usuario, sin inventarlo si no lo tiene.

    `get_profile` crea uno sobre la marcha, que es lo que hace falta para quien
    esta navegando pero no para el usuario sobre el que se actua: fabricarle un
    perfil solo para poder compararlo seria darle un rol que nadie le dio.
    """
    return UserProfile.objects.filter(user=user).first()


def _incoherencia_rol_cliente(role, cid):
    """
    Por que un usuario no puede llevar a la vez un cliente y un rol de la casa.

    El formulario ofrece el rol y el cliente como dos campos sueltos, asi que
    se podia dar de alta a alguien "para un cliente" y dejarle el rol `staff`
    que viene por omision. El perfil quedaba con cliente y con rol de operador,
    y `customer_ops_filter` solo acota a quien tiene rol 'customer': ese
    usuario veia **todas las operaciones de la empresa**, incluidas las de los
    demas clientes, mientras quien lo creo pensaba que le habia dado un acceso
    limitado. No era un fallo del filtro sino de la pantalla que permitia
    describir mal a una persona.

    Al reves tambien importa, aunque no abra nada: un 'customer' sin cliente no
    alcanza ninguna operacion -- es el fail-closed de `customer_ops_filter` --,
    asi que se le entrega a alguien una cuenta que no le sirve para nada y el
    "no veo mis operaciones" tarda dias en llegar.

    Devuelve el texto del rechazo, o cadena vacia si la combinacion es buena.
    """
    if cid and role != 'customer':
        return (f'A user linked to a customer must have the role "customer". '
                f'"{role}" is a company role and would see every operation of '
                f'every customer. Nothing was changed.')
    if role == 'customer' and not cid:
        return ('A customer user needs the customer it belongs to; without it '
                'the account cannot see a single operation. Nothing was changed.')
    return ''


@login_required
def user_management(request):
    tenant  = get_tenant_or_404(request)
    profile = get_profile(request.user)
    if not profile.can_manage_users():
        return HttpResponse('Permission denied.', status=403)
    users    = User.objects.filter(profile__tenant=tenant).order_by('username')
    profiles = {p.user_id: p for p in UserProfile.objects.select_related('customer').filter(tenant=tenant)}
    customers = Catalog.objects.filter(category='CUSTOMER', active=True, tenant=tenant).order_by('name')
    msg = ''
    # El template pintaba cualquier msg como exito, con su palomita, incluso los
    # rechazos. Con esto los errores se ven como errores.
    msg_is_error = False
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            uname = request.POST.get('username','').strip()
            pwd   = request.POST.get('password','').strip()
            role  = request.POST.get('role','staff')
            cid   = request.POST.get('customer_id','').strip()
            # El formulario pinta el campo desde siempre y el alta lo tiraba.
            # Ahora importa de verdad: sin contrasena de borrado configurada,
            # el usuario no puede borrar nada.
            del_pwd = request.POST.get('delete_password','').strip()
            if uname and pwd and not profile.can_assign_role(role):
                # El rol llegaba del formulario y se guardaba tal cual, asi que
                # un administrador podia nombrar un 'superadmin' y quedar por
                # debajo de alguien a quien acababa de crear.
                msg = f'You cannot create a user with the role "{role}".'
                msg_is_error = True
            elif uname and pwd and _incoherencia_rol_cliente(role, cid):
                msg = _incoherencia_rol_cliente(role, cid)
                msg_is_error = True
            elif uname and pwd:
                if not User.objects.filter(username=uname).exists():
                    u = User.objects.create_user(username=uname, password=pwd)
                    cat = Catalog.objects.filter(pk=int(cid), tenant=tenant).first() if cid else None
                    nuevo_perfil = UserProfile.objects.create(
                        user=u, role=role, customer=cat, tenant=tenant)
                    if del_pwd:
                        nuevo_perfil.set_delete_password(del_pwd)
                        nuevo_perfil.save(update_fields=['delete_password'])
                    # La contrasena ya no queda guardada en claro, asi que esta
                    # pantalla no va a poder recordarla: se dice aqui, una vez.
                    msg = (f'User "{uname}" created with role "{role}". '
                           f'The password is not stored anywhere — write it down '
                           f'or set a new one later.')
                else:
                    msg = f'Username "{uname}" already exists.'
                    msg_is_error = True
        elif action == 'create_customer':
            # Alta del cliente nivel 2 en un solo paso: el Catalog CUSTOMER y el
            # usuario que lo va a usar se crean juntos. Antes eran dos pantallas
            # y era facil terminar con un usuario 'customer' sin cliente asignado,
            # que es justo el caso que customer_ops_filter bloquea por completo.
            cname  = request.POST.get('customer_name', '').strip()
            uname  = request.POST.get('username', '').strip()
            pwd    = request.POST.get('password', '').strip()

            existing = Catalog.objects.filter(
                tenant=tenant, category='CUSTOMER', active=True, name__iexact=cname).first()

            if not (cname and uname and pwd):
                msg = 'Customer name, username and password are all required.'
                msg_is_error = True
            elif existing:
                msg = (f'Customer "{existing.name}" already exists. Use "Create New User" '
                       f'above and link the new user to it. Nothing was created.')
                msg_is_error = True
            elif User.objects.filter(username=uname).exists():
                msg = (f'Username "{uname}" is already taken. Nothing was created — '
                       f'pick a different username.')
                msg_is_error = True
            else:
                # Atomico a proposito: si algo falla a mitad, no queremos dejar el
                # cliente sin usuario ni el usuario sin cliente.
                with transaction.atomic():
                    cat = Catalog.objects.create(
                        tenant=tenant, category='CUSTOMER', name=cname,
                        abbreviation=request.POST.get('abbreviation', '').strip().upper() or None,
                        contact_email=request.POST.get('contact_email', '').strip() or None,
                        phone=request.POST.get('phone', '').strip() or None,
                        whatsapp=request.POST.get('whatsapp', '').strip() or None,
                    )
                    u = User.objects.create_user(username=uname, password=pwd)
                    UserProfile.objects.create(
                        user=u, tenant=tenant, role='customer',
                        customer=cat,
                    )
                msg = (f'Customer "{cat.name}" created, with login "{uname}" '
                       f'already linked to it. The password is not stored '
                       f'anywhere — write it down or set a new one later.')
        elif action == 'delete':
            uid = request.POST.get('user_id')
            u   = get_object_or_404(User, pk=uid, profile__tenant=tenant)
            if not profile.can_manage_user(_perfil_de(u)):
                msg = f'You cannot manage the account of "{u.username}".'
                msg_is_error = True
            elif u != request.user:
                u.delete()
                msg = 'User deleted.'
        elif action == 'change_password':
            uid = request.POST.get('user_id')
            new_pwd = request.POST.get('new_password','').strip()
            if uid and new_pwd:
                u = get_object_or_404(User, pk=uid, profile__tenant=tenant)
                # Validar solo el rol dejaba la puerta entornada: sin esto un
                # administrador no podia nombrar un superadmin, pero si cambiarle
                # la contrasena al que ya hubiera y entrar como el.
                if not profile.can_manage_user(_perfil_de(u)):
                    msg = f'You cannot change the password of "{u.username}".'
                    msg_is_error = True
                else:
                    # Solo la de Django, cifrada. Antes se guardaba ademas una
                    # copia en claro en el perfil para poder repintarla en la
                    # tabla; el precio era tener todas las contrasenas de la
                    # empresa legibles en la base.
                    u.set_password(new_pwd)
                    u.save()
                    msg = (f'Password updated for "{u.username}". It is not stored '
                           f'anywhere — hand it over now.')
        elif action == 'set_delete_password':
            # Tiene accion propia, y no el formulario de edicion, porque la fila
            # ofrecia dos campos de contrasena muy distintos: el ancho de la
            # izquierda cambia la contrasena de acceso y el estrecho del final
            # la de borrado. Quien queria lo segundo tenia buenas razones para
            # escribir en lo primero, y el efecto era cambiarle a alguien la
            # contrasena con la que entra.
            uid = request.POST.get('user_id')
            nueva = request.POST.get('delete_password', '').strip()
            u = get_object_or_404(User, pk=uid, profile__tenant=tenant)
            objetivo = _perfil_de(u)
            if not profile.can_manage_user(objetivo):
                msg = f'You cannot change the delete password of "{u.username}".'
                msg_is_error = True
            elif objetivo is None:
                msg = f'"{u.username}" has no profile in this company.'
                msg_is_error = True
            else:
                objetivo.set_delete_password(nueva)
                objetivo.save(update_fields=['delete_password'])
                msg = (f'Delete password set for "{u.username}".' if nueva
                       else f'Delete password removed for "{u.username}" — '
                            f'they can no longer delete anything.')
        elif action == 'update_role':
            uid  = request.POST.get('user_id')
            role = request.POST.get('role','staff')
            cid  = request.POST.get('customer_id','').strip()
            del_pwd = request.POST.get('delete_password','').strip()
            u    = get_object_or_404(User, pk=uid, profile__tenant=tenant)
            if not profile.can_manage_user(_perfil_de(u)) or not profile.can_assign_role(role):
                msg = f'You cannot assign the role "{role}" to "{u.username}".'
                msg_is_error = True
            elif _incoherencia_rol_cliente(role, cid):
                msg = _incoherencia_rol_cliente(role, cid)
                msg_is_error = True
            else:
                p, _ = UserProfile.objects.get_or_create(user=u, defaults={'tenant': tenant})
                p.role     = role
                p.customer = Catalog.objects.filter(pk=int(cid), tenant=tenant).first() if cid else None
                if del_pwd:
                    # Cifrada: es el control que autoriza a borrar, y estuvo
                    # guardada en claro y a la vista en esta misma pantalla.
                    p.set_delete_password(del_pwd)
                p.save()
                msg = f'User "{u.username}" updated.'
        users    = User.objects.filter(profile__tenant=tenant).order_by('username')
        profiles = {p.user_id: p for p in UserProfile.objects.select_related('customer').filter(tenant=tenant)}

    users    = User.objects.filter(profile__tenant=tenant).order_by('username')
    profiles = {p.user_id: p for p in UserProfile.objects.select_related('customer').filter(tenant=tenant)}
    # Se recalcula despues del POST para que incluya un cliente recien creado.
    customers = Catalog.objects.filter(category='CUSTOMER', active=True, tenant=tenant).order_by('name')
    return render(request, 'warehouse/partials/user_management.html', {
        'users': users, 'profiles': profiles,
        'customers': customers, 'msg': msg, 'msg_is_error': msg_is_error,
        'request': request,
        # Para no ofrecer en el desplegable un rol que la vista va a rechazar.
        'profile': profile,
    })


# ── DEBUG ─────────────────────────────────────────────────────────────────────

@login_required
def debug_catalog(request):
    tenant = get_tenant_or_404(request)
    def to_list(qs):
        return [{'id': e.pk, 'name': e.name} for e in qs]
    return JsonResponse({
        'customers':    to_list(Catalog.objects.filter(category='CUSTOMER',    active=True, tenant=tenant).order_by('name')),
        'shippers':     to_list(Catalog.objects.filter(category='SHIPPER',     active=True, tenant=tenant).order_by('name')),
        'carriers':     to_list(Catalog.objects.filter(category='CARRIER',     active=True, tenant=tenant).order_by('name')),
        'bundle_types': to_list(Catalog.objects.filter(category='BUNDLE_TYPE', active=True, tenant=tenant).order_by('name')),
    })


# ── MOBILE ────────────────────────────────────────────────────────────────────

@login_required
def mobile_dashboard(request):
    tenant = get_tenant_or_404(request)
    profile = get_profile(request.user)
    ops = WarehouseOperation.objects.filter(tenant=tenant).select_related(
        'customer','shipper','carrier','bundle_type')
    ops = customer_ops_filter(request.user, ops)[:200]
    def cat_json(category):
        return json.dumps([
            {'id': e.pk, 'name': e.name}
            for e in Catalog.objects.filter(category=category, active=True, tenant=tenant).order_by('name')
        ])
    context = {
        'profile': profile,
        'is_home': profile.is_home(),
        'tenant': tenant,
        'operations': ops,
        'operational_entries': _catalog_entries(tenant, 'operational'),
        'customer_entries':    _catalog_entries(tenant, 'customers'),
        'can_edit_operational': profile.can_edit_catalog('SHIPPER'),
        'can_edit_customers':   profile.can_edit_catalog('CUSTOMER'),
        'customers_json':    cat_json('CUSTOMER'),
        'shippers_json':     cat_json('SHIPPER'),
        'carriers_json':     cat_json('CARRIER'),
        'bundle_types_json': cat_json('BUNDLE_TYPE'),
    }
    return render(request, 'warehouse/mobile.html', context)


@login_required
@require_GET
def exit_entry_totals(request):
    tenant = get_tenant_or_404(request) #####072526 13:28 3.7. exit_entry_totals
    ids_str = request.GET.get('ids', '')
    if not ids_str:
        return JsonResponse({})
    custom_ids = [x.strip() for x in ids_str.split(',') if x.strip()]
    entries = WarehouseOperation.objects.filter(
        tenant=tenant, #####072526 13:28 3.7. exit_entry_totals
        custom_id__in=custom_ids, operation_type='ENTRY'
    ).select_related('shipper', 'bundle_type')

    total_qty  = 0
    total_lbs  = 0
    total_kgs  = 0
    desc_parts = []
    po_parts   = []
    shippers   = set()
    bundle_types = set()

    for e in entries:
        total_qty += e.bundle_qty or 0
        total_lbs += float(e.weight_lbs or 0)
        total_kgs += float(e.weight_kgs or 0)
        if e.description:
            desc_parts.append(e.description)
        if e.po_order:
            po_parts.append(e.po_order)
        s = e.get_shipper_display()
        if s and s != '—': shippers.add(s)
        bt = e.get_bundle_type_display_name()
        if bt and bt != '—': bundle_types.add(bt)

    shipper_val     = list(shippers)[0] if len(shippers) == 1 else ('VARIOUS' if shippers else '')
    bundle_type_val = ', '.join(bundle_types) if bundle_types else ''

    return JsonResponse({
        'bundle_qty':   total_qty,
        'weight_lbs':   round(total_lbs, 2),
        'weight_kgs':   round(total_kgs, 2),
        'description':  ' / '.join(desc_parts),
        'po_order':     ' / '.join(po_parts),
        'shipper':      shipper_val,
        'bundle_type':  bundle_type_val,
    })


@login_required
def operation_edit(request, pk):
    tenant = get_tenant_or_404(request) #### Usa tenant=tenant en get_object_or_404 072526 21:05
    op = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant) ##### 5.5. operation_edit 072526 20:15
    profile = get_profile(request.user)
    # El staff tambien corrige: exigir `is_home()` dejaba al operador que
    # captura sin forma de arreglar un peso mal tecleado, y ademas el boton
    # Edit le aparecia en la tabla y respondia 403.
    if not profile.can_edit_operations() and not profile.is_customer():
        return HttpResponse('Permission denied.', status=403)
    if not customer_can_access_op(request.user, op):
        return HttpResponse('Permission denied.', status=403)

    if request.method == 'POST':
        p = request.POST
        def to_dec(v):
            try:
                from decimal import Decimal
                return Decimal(v) if v and str(v).strip() else None
            except: return None
        def to_int(v):
            try: return int(v) if v and str(v).strip() else None
            except: return None

        if profile.is_customer():
            op.customer_notes = p.get('customer_notes', op.customer_notes or '')
            op.save(update_fields=['customer_notes', 'updated_at'])
        else:
            op.date             = p.get('date', str(op.date))
            op.entry_dispatched = p.get('entry_dispatched', '')
            op.invoice          = p.get('invoice', '')
            op.po_order         = p.get('po_order', '')
            op.seal             = p.get('seal', '')
            op.pro              = p.get('pro', '')
            op.trailer          = p.get('trailer', '')
            op.bundle_qty       = to_int(p.get('bundle_qty'))
            op.weight_lbs       = to_dec(p.get('weight_lbs'))
            op.weight_kgs       = to_dec(p.get('weight_kgs'))
            op.description      = p.get('description', '')
            op.note             = p.get('note', '')
            op.damage           = p.get('damage') == '1'
            op.damage_description = p.get('damage_description', '')
            op.customer_notes   = p.get('customer_notes', op.customer_notes or '')
            op.customer_name_manual  = p.get('customer_name_manual', op.customer_name_manual or '')
            op.shipper_name_manual   = p.get('shipper_name_manual',  op.shipper_name_manual  or '')
            op.carrier_name_manual   = p.get('carrier_name_manual',  op.carrier_name_manual  or '')
            op.bundle_type_manual    = p.get('bundle_type_manual',   op.bundle_type_manual   or '')
            # Nuevos campos
            op.ref_aa = p.get('ref_aa', op.ref_aa or '')
            op.ref_dys = p.get('ref_dys', op.ref_dys or '')
            op.pedimento = p.get('pedimento', op.pedimento or '')
            op.save()

        if request.headers.get('HX-Request'):
            return HttpResponse(
                '<script>'
                'document.getElementById("modal")?.classList.remove("on");'
                'if(typeof htmx !== "undefined") htmx.ajax("GET", "/operations/search/", {target:"#ops-table", swap:"innerHTML"});'
                '</script>'
            )
        else:
            return HttpResponse(
                '<script>'
                'if(window.opener){'
                '   if(typeof window.opener.htmx !== "undefined"){'
                '       window.opener.htmx.ajax("GET", "/operations/search/", {target:"#ops-table", swap:"innerHTML"});'
                '   } else {'
                '       window.opener.location.reload();'
                '   }'
                '   window.close();'
                '} else {'
                '   window.location.href = "/dashboard/";'
                '}'
                '</script>'
            )

    edit_fields = [
        ('Date',             'date',             str(op.date)),
        ('Type',             'operation_type',   op.operation_type),
        ('Entries Disp.',    'entry_dispatched',  op.entry_dispatched),
        ('Customer',         'customer_name_manual', op.get_customer_display()),
        ('Shipper',          'shipper_name_manual',  op.get_shipper_display()),
        ('Invoice',          'invoice',          op.invoice),
        ('PO / Order',       'po_order',         op.po_order),
        ('Seal',             'seal',             op.seal),
        ('Carrier',          'carrier_name_manual',  op.get_carrier_display()),
        ('PRO',              'pro',              op.pro),
        ('Trailer',          'trailer',          op.trailer),
        ('Bundle Type',      'bundle_type_manual',   op.get_bundle_type_display_name()),
        ('Bundle Qty',       'bundle_qty',       op.bundle_qty),
        ('Weight LBS',       'weight_lbs',       op.weight_lbs),
        ('Weight KGS',       'weight_kgs',       op.weight_kgs),
        ('Description',      'description',      op.description),
        ('Note',             'note',             op.note),
        ('Damage Desc.',     'damage_description', op.damage_description),
        ('REF AA',           'ref_aa',           op.ref_aa or ''),
        ('DYS',              'ref_dys',          op.ref_dys or ''),
        ('PEDIMENTO',        'pedimento',        op.pedimento or ''),
    ]
    return render(request, 'warehouse/partials/operation_edit.html', {
        'operation': op, 'edit_fields': edit_fields,
        'profile': profile,
    })


@login_required
def report_generator_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl not installed. Run: pip install openpyxl', status=500)

    tenant = get_tenant_or_404(request)
    ops_ids = request.GET.get('ids','').split(',')
    ops = WarehouseOperation.objects.filter(pk__in=[i for i in ops_ids if i], tenant=tenant).select_related(
        'customer','shipper','carrier','bundle_type').order_by('-date')
    ops = customer_ops_filter(request.user, ops)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Operations Report'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F172A')
    headers = ['#','Date','Custom ID','Type','Status','Customer','Shipper','Carrier',
               'Invoice','PO/Order','Seal','PRO','Trailer','Bundle Type',
               'Bundle Qty','Weight LBS','Weight KGS','Description','Note',
               'Entries Dispatched','Damage','Email Sent', 'REF AA', 'DYS', 'PEDIMENTO']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, op in enumerate(ops, 2):
        ws.append([
            op.pk, str(op.date), op.custom_id, op.get_operation_type_display(),
            op.status,
            op.get_customer_display(), op.get_shipper_display(), op.get_carrier_display(),
            op.invoice or '', op.po_order or '', op.seal or '',
            op.pro or '', op.trailer or '', op.get_bundle_type_display_name(),
            op.bundle_qty, float(op.weight_lbs) if op.weight_lbs else None,
            float(op.weight_kgs) if op.weight_kgs else None,
            op.description or '', op.note or '', op.entry_dispatched or '',
            'Yes' if op.damage else 'No', 'Yes' if op.email_sent else 'No',
            op.ref_aa or '', op.ref_dys or '', op.pedimento or '',
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="operations_report.xlsx"'
    return resp


# ── IMPORT / EXPORT — OPERATIONS ─────────────────────────────────────────────

@login_required
def operations_layout(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Operations Import'

    headers = [
        'date (YYYY-MM-DD)', 'operation_type (ENTRY/EXIT)', 'customer_name',
        'shipper_name', 'carrier_name', 'bundle_type_name',
        'invoice', 'po_order', 'seal', 'pro', 'trailer',
        'bundle_qty', 'weight_lbs', 'weight_kgs',
        'description', 'note', 'damage (yes/no)', 'damage_description',
        'entry_dispatched', 'ref_aa', 'ref_dys', 'pedimento',
    ]
    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='0F172A')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)

    ws.append([
        '2026-04-01', 'ENTRY', 'ACME Corp', 'Fast Ship', 'SuperCarrier', 'PALLET',
        'INV-001', 'PO-001', 'SEAL-001', 'PRO-001', 'TRAIL-001',
        10, 500.00, 226.80,
        'Electronic components', 'Handle with care', 'no', '',
        '', '', '', '',
    ])
    ws['A2'].font = Font(italic=True, color='94a3b8')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="operations_import_layout.xlsx"'
    return resp


@login_required
@require_POST
def operations_import(request):
    tenant = get_tenant_or_404(request)
    profile = get_profile(request.user)
    if not profile.can_create_operations():
        return HttpResponse('Permission denied.', status=403)
    try:
        import openpyxl
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    f = request.FILES.get('import_file')
    if not f:
        return HttpResponse('No file uploaded.', status=400)

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        created = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not any(row):
                continue
            try:
                date_val, op_type, cust_name, shipper_name, carrier_name, bundle_name, \
                invoice, po_order, seal, pro, trailer, \
                bundle_qty, weight_lbs, weight_kgs, \
                description, note, damage_str, damage_desc, entry_disp, \
                ref_aa, ref_dys, pedimento = (list(row) + [None]*25)[:22]

                if not date_val or not op_type:
                    continue

                from datetime import date as date_cls
                if hasattr(date_val, 'date'):
                    op_date = date_val.date()
                elif isinstance(date_val, date_cls):
                    op_date = date_val
                else:
                    from datetime import datetime
                    op_date = datetime.strptime(str(date_val), '%Y-%m-%d').date()

                op_type = str(op_type).strip().upper()
                if op_type not in ('ENTRY', 'EXIT'):
                    errors.append(f'Row {row_idx}: invalid type "{op_type}"')
                    continue

                def find_cat(name, category):
                    if not name: return None, ''
                    name = str(name).strip()
                    obj = Catalog.objects.filter(category=category, name__iexact=name, active=True, tenant=tenant).first()
                    return (obj, '') if obj else (None, name)

                cust_obj, cust_manual       = find_cat(cust_name, 'CUSTOMER')
                ship_obj, ship_manual       = find_cat(shipper_name, 'SHIPPER')
                carr_obj, carr_manual       = find_cat(carrier_name, 'CARRIER')
                bundle_obj, bundle_manual   = find_cat(bundle_name, 'BUNDLE_TYPE')

                from decimal import Decimal
                def to_dec(v):
                    try: return Decimal(str(v)) if v is not None and str(v).strip() else None
                    except: return None
                def to_int(v):
                    try: return int(v) if v is not None else None
                    except: return None

                op = WarehouseOperation(
                    tenant=tenant,
                    date=op_date, operation_type=op_type,
                    customer=cust_obj, customer_name_manual=cust_manual,
                    shipper=ship_obj, shipper_name_manual=ship_manual,
                    carrier=carr_obj, carrier_name_manual=carr_manual,
                    bundle_type=bundle_obj, bundle_type_manual=bundle_manual,
                    invoice=str(invoice or '').strip(),
                    po_order=str(po_order or '').strip(),
                    seal=str(seal or '').strip(),
                    pro=str(pro or '').strip(),
                    trailer=str(trailer or '').strip(),
                    bundle_qty=to_int(bundle_qty),
                    weight_lbs=to_dec(weight_lbs),
                    weight_kgs=to_dec(weight_kgs),
                    description=str(description or '').strip(),
                    note=str(note or '').strip(),
                    damage=str(damage_str or '').strip().lower() == 'yes',
                    damage_description=str(damage_desc or '').strip(),
                    entry_dispatched=str(entry_disp or '').strip(),
                    created_by=request.user,
                    ref_aa=str(ref_aa or '').strip(),
                    ref_dys=str(ref_dys or '').strip(),
                    pedimento=str(pedimento or '').strip(),
                )
                op.save()
                created += 1
            except Exception as e:
                errors.append(f'Row {row_idx}: {e}')

        ops = WarehouseOperation.objects.filter(tenant=tenant).select_related(
            'customer', 'shipper', 'carrier', 'bundle_type')
        ops = customer_ops_filter(request.user, ops)[:200]
        profile = get_profile(request.user)
        msg = f'{created} operation(s) imported successfully.'
        if errors:
            msg += f' Errors: {"; ".join(errors[:3])}'
        return render(request, 'warehouse/partials/operations_table.html', {
            'operations': ops,
            'import_success' if not errors else 'import_error': msg,
            'is_home': is_home(request.user), 'profile': profile,
        })
    except Exception as e:
        return HttpResponse(f'Import failed: {e}', status=500)


# ── IMPORT / EXPORT — CATALOG ─────────────────────────────────────────────────

@login_required
def catalog_layout(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Catalog Import'

    headers = ['category', 'name', 'contact_email', 'phone', 'whatsapp', 'address', 'notes']
    notes_row = [
        'CUSTOMER / SHIPPER / CARRIER / BUNDLE_TYPE / CC_EMAIL',
        'Required', 'Optional, comma-separated', 'Optional',
        'Optional (+521XXXXXXXXXX)', 'Optional', 'Optional',
    ]
    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='0F172A')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h)+4, 20)

    ws.append(notes_row)
    for cell in ws[2]:
        cell.font = Font(italic=True, color='94a3b8')

    ws.append(['CUSTOMER', 'Example Corp', 'contact@example.com', '+5218001234567', '+5218001234567', '123 Main St', 'VIP client'])

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="catalog_import_layout.xlsx"'
    return resp


@login_required
@require_POST
def catalog_import(request):
    tenant = get_tenant_or_404(request)
    profile = get_profile(request.user)
    if profile.is_customer():
        return HttpResponse('Permission denied.', status=403)
    try:
        import openpyxl
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    scope = _catalog_scope(request)
    f = request.FILES.get('import_file')
    if not f:
        return HttpResponse('No file.', status=400)

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        created = 0
        errors = []
        valid_cats = ['CUSTOMER','SHIPPER','CARRIER','BUNDLE_TYPE','TYPE_OP','CC_EMAIL']

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not any(row): continue
            try:
                category, name, email, phone, whatsapp, address, notes = (list(row)+[None]*7)[:7]
                category = str(category or '').strip().upper()
                name     = str(name or '').strip()
                if not category or not name or category not in valid_cats:
                    errors.append(f'Row {row_idx}: invalid category or missing name')
                    continue
                # El Excel es la puerta de atrás del catálogo: sin esto, quien no
                # puede dar de alta un cliente desde el formulario lo da de alta
                # subiendo un archivo con la categoría escrita dentro.
                if not profile.can_edit_catalog(category):
                    errors.append(
                        f'Row {row_idx}: only an administrator can import {category} entries')
                    continue
                if not Catalog.objects.filter(category=category, name__iexact=name, tenant=tenant).exists():
                    Catalog.objects.create(
                        tenant=tenant,
                        category=category, name=name,
                        contact_email=str(email or '').strip(),
                        phone=str(phone or '').strip(),
                        whatsapp=str(whatsapp or '').strip(),
                        address=str(address or '').strip(),
                        notes=str(notes or '').strip(),
                    )
                    created += 1
            except Exception as e:
                errors.append(f'Row {row_idx}: {e}')

        msg = f'{created} catalog entries imported.'
        if errors: msg += f' Errors: {"; ".join(errors[:3])}'
        key = 'import_success' if not errors else 'import_error'
        return render(request, 'warehouse/partials/catalog_table.html',
                      _catalog_table_context(request, tenant, scope, **{key: msg}))
    except Exception as e:
        return HttpResponse(f'Import failed: {e}', status=500)

# ── SAAS PLATFORM (nivel 1: fuera del scoping de tenant) ─────────────────────
#
# Este nivel no pasa por `get_tenant_or_404`: sus usuarios no pertenecen a
# ninguna empresa. El acceso se resuelve con `platform_role`, que devuelve
# 'admin', 'staff' o None.

def platform_role(user):
    """
    Nivel de plataforma de este usuario, o None si no tiene ninguno.

    El `is_superuser` de Django todavia abre esta puerta, pero solo mientras no
    haya un administrador de plataforma de verdad. En cuanto existe el primer
    `PlatformUser` con rol 'admin', la llave maestra deja de valer aqui y el
    acceso pasa a ser el que este modelo dice.

    La condicion es la que resuelve el huevo y la gallina sin poder dejar a
    nadie fuera: la pantalla que reparte este acceso solo la ve quien ya lo
    tiene, asi que hace falta una llave para crear al primero; y esa llave se
    retira sola en cuanto hay sucesor, en vez de quedarse ahi para siempre
    esperando a que alguien se acuerde. Si el sucesor se revoca, vuelve.
    """
    if not user.is_authenticated:
        return None
    acceso = PlatformUser.objects.filter(user=user).first()
    if acceso:
        return acceso.role
    if user.is_superuser and not PlatformUser.objects.filter(role='admin').exists():
        return 'admin'
    return None


def _es_admin_de_plataforma(user):
    return platform_role(user) == 'admin'


def _sin_permiso_de_plataforma(user, solo_admin=False):
    """
    Devuelve la respuesta de rechazo, o None si puede pasar.

    Se comprueban las dos cosas por separado porque casi todas las pantallas de
    plataforma las ve el soporte y solo algunas acciones son del administrador.
    """
    nivel = platform_role(user)
    if nivel is None:
        return HttpResponse('Permission denied.', status=403)
    if solo_admin and nivel != 'admin':
        return HttpResponse('Only a platform administrator can do this.', status=403)
    return None



# Formatos que reportlab sabe dibujar en un PDF y que un navegador enseña. El
# SVG queda fuera a proposito: reportlab no lo pinta sin una libreria aparte, y
# aceptarlo aqui daria un logo que se ve en la pantalla y desaparece en el
# documento, que es donde hace falta.
FORMATOS_DE_LOGO = ('.png', '.jpg', '.jpeg', '.gif')

# Tope del archivo. Un logo es una imagen pequeña; lo que llega de mas suele ser
# una foto sin recortar, y acaba en cada PDF que la empresa manda.
MAX_LOGO_MB = 2


def _guardar_logo(tenant, archivo):
    """
    Guarda el logo de una empresa. Devuelve el error, o cadena vacia si fue bien.

    Devuelve el problema en vez de lanzarlo porque el alta de una empresa no
    puede fallar entera por el logo: si la imagen no sirve, la empresa queda
    creada -- con su administrador y su suscripcion -- y lo que se avisa es que
    el logo no se puso.
    """
    if archivo is None:
        return ''

    nombre = (archivo.name or '').lower()
    if not nombre.endswith(FORMATOS_DE_LOGO):
        return (f'El logo no se guardo: tiene que ser '
                f'{", ".join(FORMATOS_DE_LOGO)}.')
    if archivo.size > MAX_LOGO_MB * 1024 * 1024:
        return f'El logo no se guardo: pasa de {MAX_LOGO_MB} MB.'

    anterior = tenant.logo.name if tenant.logo else ''
    try:
        tenant.logo = archivo
        tenant.save(update_fields=['logo'])
    except Exception as e:
        logger.warning('No se pudo guardar el logo de %s: %s', tenant.name, e)
        return 'El logo no se guardo: el almacen no respondio.'

    # El anterior deja de estar referenciado por nadie: si se queda, el bucket
    # acumula un logo por cada cambio y ninguno se puede distinguir del vivo.
    if anterior and anterior != tenant.logo.name:
        try:
            tenant.logo.storage.delete(anterior)
        except Exception as e:
            logger.warning('Quedo un logo huerfano en %s: %s', anterior, e)
    return ''


@login_required
def platform_tenant_list(request):
    negado = _sin_permiso_de_plataforma(request.user)
    if negado:
        return negado
    es_admin = _es_admin_de_plataforma(request.user)

    msg = ''
    if request.method == 'POST':
        # Dar de alta una empresa es fabricar la llave de su administrador, y
        # desactivarla es cortarle el servicio: las dos son del administrador de
        # plataforma, no del soporte.
        if not es_admin:
            return HttpResponse('Only a platform administrator can do this.', status=403)
        action = request.POST.get('action')

        if action == 'create':
            name          = request.POST.get('name', '').strip()
            subdomain_in  = request.POST.get('subdomain', '').strip()
            plan          = request.POST.get('plan', 'starter')
            billing_email = request.POST.get('billing_email', '').strip()
            admin_username = request.POST.get('admin_username', '').strip()
            admin_password = request.POST.get('admin_password', '').strip()

            subdomain = re.sub(r'[^a-z0-9-]', '', slugify(subdomain_in or name))
            if not name:
                msg = 'Tenant name is required.'
            elif not subdomain:
                msg = 'Could not derive a valid subdomain from the name provided.'
            elif Tenant.objects.filter(subdomain=subdomain).exists():
                msg = f'Subdomain "{subdomain}" is already in use.'
            elif admin_username and User.objects.filter(username=admin_username).exists():
                msg = f'Username "{admin_username}" is already taken. Tenant was NOT created — pick a different admin username.'
            else:
                tenant = Tenant.objects.create(
                    name=name, type='organization', subdomain=subdomain,
                    is_active=True, plan=plan, billing_email=billing_email or None,
                )
                # El logo sale en los reportes y etiquetas que la empresa manda
                # a sus clientes. Es opcional: sin el, los documentos llevan su
                # nombre en texto.
                error_logo = _guardar_logo(tenant, request.FILES.get('logo'))
                Subscription.objects.create(tenant=tenant, plan=plan)
                msg = f'Tenant "{name}" created (subdomain: {subdomain}).'
                # El aviso del logo va detras del de creacion y no delante: la
                # empresa quedo creada, que es lo primero que hay que leer.
                if error_logo:
                    msg += ' ' + error_logo

                if admin_username and admin_password:
                    admin_user = User.objects.create_user(username=admin_username, password=admin_password)
                    UserProfile.objects.create(
                        user=admin_user, tenant=tenant, role='admin',
                    )
                    msg += (f' Admin user "{admin_username}" created for this tenant. '
                            f'Its password is not stored anywhere — hand it over now.')

        elif action == 'set_logo':
            t = get_object_or_404(Tenant, pk=request.POST.get('tenant_id'))
            archivo = request.FILES.get('logo')
            if archivo is None and request.POST.get('quitar') == '1':
                if t.logo:
                    t.logo.delete(save=False)
                t.logo = None
                t.save(update_fields=['logo'])
                msg = f'Se quito el logo de "{t.name}".'
            else:
                error_logo = _guardar_logo(t, archivo)
                msg = error_logo or f'Logo de "{t.name}" actualizado.'

        elif action == 'toggle_active':
            tid = request.POST.get('tenant_id')
            t = get_object_or_404(Tenant, pk=tid)
            t.is_active = not t.is_active
            t.save(update_fields=['is_active'])
            msg = f'Tenant "{t.name}" is now {"active" if t.is_active else "inactive"}.'

    tenants = Tenant.objects.filter(type='organization').annotate(
        user_count=Count('users', distinct=True),
        op_count=Count('operations', distinct=True),
    ).order_by('name')

    return render(request, 'warehouse/partials/platform_tenants.html', {
        'tenants': tenants, 'msg': msg,
        'platform_role': platform_role(request.user),
        'is_platform_admin': es_admin,
    })


@login_required
def platform_notifications(request):
    """
    La bitácora de envíos, para soporte.

    Es el trabajo diario del staff de plataforma y la razón principal por la que
    ese nivel existe: el caso real es «a este cliente no le llegan los correos»,
    y `NotificationLog` ya guarda el estado, el motivo y el destinatario de cada
    intento. Hasta ahora solo se veía desde el admin de Django, que exige
    `is_superuser`; o sea que para leer una bitácora había que entregar las
    llaves del sistema entero.

    Es de solo lectura para los dos niveles. No hay acciones: reenviar un aviso
    es cosa del operador de la empresa, desde el detalle de su operación.
    """
    negado = _sin_permiso_de_plataforma(request.user)
    if negado:
        return negado

    registros = (NotificationLog.objects
                 .select_related('tenant', 'customer', 'triggered_by')
                 .order_by('-created_at'))

    tenant_id = (request.GET.get('tenant') or '').strip()
    estado    = (request.GET.get('status') or '').strip()
    if tenant_id.isdigit():
        registros = registros.filter(tenant_id=int(tenant_id))
    if estado:
        registros = registros.filter(status=estado)

    return render(request, 'warehouse/partials/platform_notifications.html', {
        'registros': registros[:200],
        'tenants': Tenant.objects.order_by('name'),
        'tenant_id': tenant_id,
        'estado': estado,
        'estados': NotificationLog.STATUS_CHOICES,
        'platform_role': platform_role(request.user),
    })


@login_required
def platform_invoices(request):
    """
    Facturacion de la plataforma: quien pago, quien debe y quien va tarde.

    Lo que habia antes eran tres campos sueltos dentro de `Subscription`, que
    es una fila por empresa: cabia una sola factura por cliente y emitir la del
    mes borraba la del mes anterior. Esta pantalla es la primera vez que se
    puede facturar de verdad.

    **El soporte mira y el administrador actua.** El staff de plataforma ve el
    listado completo y el estado de cada empresa -que es lo que necesita para
    responder «¿este cliente esta al corriente?»- pero emitir, cobrar y
    cancelar son del administrador. Cada POST lo vuelve a comprobar: esconder
    los botones no es un permiso.
    """
    negado = _sin_permiso_de_plataforma(request.user)
    if negado:
        return negado
    es_admin = _es_admin_de_plataforma(request.user)

    msg = error = ''
    if request.method == 'POST':
        negado = _sin_permiso_de_plataforma(request.user, solo_admin=True)
        if negado:
            return negado
        msg, error = _accion_de_facturacion(request)

    facturas = Invoice.objects.select_related('tenant', 'emitida_por')

    tenant_id = (request.GET.get('tenant') or '').strip()
    estado    = (request.GET.get('status') or '').strip()
    if tenant_id.isdigit():
        facturas = facturas.filter(tenant_id=int(tenant_id))
    if estado == 'vencida':
        # No es un estado guardado: es pendiente con la fecha pasada.
        facturas = facturas.filter(estado=Invoice.PENDIENTE,
                                   vence_el__lt=timezone.localdate())
    elif estado:
        facturas = facturas.filter(estado=estado)

    return render(request, 'warehouse/partials/platform_invoices.html', {
        'facturas': facturas[:200],
        'resumen': _resumen_de_facturacion(),
        'tenants': Tenant.objects.filter(type='organization').order_by('name'),
        'tenant_id': tenant_id,
        'estado': estado,
        'estados': Invoice.ESTADOS,
        'hoy': timezone.localdate(),
        'vencimiento_sugerido': timezone.localdate() + timedelta(days=DIAS_DE_VENCIMIENTO),
        'mes_sugerido': timezone.localdate().strftime('%Y-%m'),
        'msg': msg, 'error': error,
        'platform_role': platform_role(request.user),
        'is_platform_admin': es_admin,
    })


def _accion_de_facturacion(request):
    """Ejecuta la accion del POST y devuelve (mensaje, error)."""
    accion = request.POST.get('action')

    if accion == 'emitir':
        return _emitir_factura(request)

    if accion == 'pagar':
        factura = get_object_or_404(Invoice, pk=request.POST.get('invoice_id'))
        try:
            factura.marcar_pagada(referencia=request.POST.get('referencia', ''))
        except ValueError as e:
            return '', str(e)
        return f'{factura.numero} marcada como pagada.', ''

    if accion == 'enviar':
        factura = get_object_or_404(Invoice, pk=request.POST.get('invoice_id'))
        if factura.estado == Invoice.CANCELADA:
            return '', 'Una factura cancelada no se le manda a nadie.'
        enviado, detalle = notifications.enviar_factura(
            factura, triggered_by=request.user)
        if not enviado:
            return '', detalle
        return f'{factura.numero} enviada a {detalle}.', ''

    if accion == 'cancelar':
        factura = get_object_or_404(Invoice, pk=request.POST.get('invoice_id'))
        motivo = (request.POST.get('motivo') or '').strip()
        if not motivo:
            # Cancelar deja el numero ocupado para siempre; que conste por que.
            return '', 'Hace falta un motivo para cancelar una factura.'
        try:
            factura.cancelar(motivo)
        except ValueError as e:
            return '', str(e)
        return f'{factura.numero} cancelada.', ''

    return '', ''


def _emitir_factura(request):
    """
    Da de alta una factura nueva.

    El numero se aparta al final y dentro de la transaccion, ya validado todo
    lo demas: un numero apartado no vuelve, asi que gastarlo para acabar
    rechazando el formulario deja un hueco en la serie sin ninguna razon.
    """
    tenant_id = (request.POST.get('tenant_id') or '').strip()
    if not tenant_id.isdigit():
        return '', 'Elige la empresa a la que se factura.'
    empresa = Tenant.objects.filter(pk=int(tenant_id), type='organization').first()
    if empresa is None:
        return '', 'Esa empresa no existe.'

    monto = _monto_valido(request.POST.get('monto'))
    if monto is None:
        return '', 'El monto tiene que ser un número mayor que cero.'

    periodo = _periodo_del_mes(request.POST.get('periodo'))
    if periodo is None:
        return '', 'El periodo no es un mes válido.'
    periodo_inicio, periodo_fin = periodo

    vence = _fecha_valida(request.POST.get('vence_el'))
    if vence is None:
        return '', 'La fecha de vencimiento no es válida.'

    ya_existe = Invoice.objects.filter(
        tenant=empresa, periodo_inicio=periodo_inicio,
        estado__in=[Invoice.PENDIENTE, Invoice.PAGADA]).first()
    if ya_existe:
        # Duplicar el mes es el error caro: el cliente recibe dos cobros por lo
        # mismo. Se avisa y se deja pasar solo si lo confirma a proposito.
        if request.POST.get('confirmar_duplicado') != '1':
            return '', (f'{empresa.name} ya tiene la factura {ya_existe.numero} '
                        f'para ese periodo. Marca la casilla si aun así quieres '
                        f'emitir otra.')

    hoy = timezone.localdate()
    with transaction.atomic():
        factura = Invoice.objects.create(
            tenant=empresa,
            numero=Invoice.siguiente_numero(hoy.year),
            periodo_inicio=periodo_inicio, periodo_fin=periodo_fin,
            emitida_el=hoy, vence_el=vence,
            plan=empresa.plan or '',
            monto_usd=monto,
            notas=(request.POST.get('notas') or '').strip(),
            emitida_por=request.user,
        )
    return f'Factura {factura.numero} emitida a {empresa.name} por {monto} USD.', ''


def _monto_valido(texto):
    try:
        monto = Decimal((texto or '').strip().replace(',', ''))
    except (InvalidOperation, AttributeError):
        return None
    if monto <= 0:
        return None
    return monto.quantize(Decimal('0.01'))


def _periodo_del_mes(texto):
    """De `2026-08` a los dos extremos del mes."""
    try:
        anio, mes = (texto or '').split('-')
        inicio = date(int(anio), int(mes), 1)
    except (ValueError, TypeError):
        return None
    dias = calendar.monthrange(inicio.year, inicio.month)[1]
    return inicio, date(inicio.year, inicio.month, dias)


def _fecha_valida(texto):
    try:
        anio, mes, dia = (texto or '').split('-')
        return date(int(anio), int(mes), int(dia))
    except (ValueError, TypeError):
        return None


def _resumen_de_facturacion():
    """
    Lo que se lee de un vistazo: cobrado, por cobrar y atrasado.

    El vencido va aparte del pendiente aunque sea un subconjunto suyo, porque
    es la unica cifra que pide hacer algo hoy.
    """
    hoy = timezone.localdate()
    pendientes = Invoice.objects.filter(estado=Invoice.PENDIENTE)
    vencidas   = pendientes.filter(vence_el__lt=hoy)
    pagadas    = Invoice.objects.filter(estado=Invoice.PAGADA,
                                        emitida_el__year=hoy.year)

    def total(qs):
        return qs.aggregate(t=Sum('monto_usd'))['t'] or Decimal('0.00')

    return {
        'pendiente_total': total(pendientes), 'pendiente_cuenta': pendientes.count(),
        'vencido_total':   total(vencidas),   'vencido_cuenta':   vencidas.count(),
        'cobrado_total':   total(pagadas),    'cobrado_cuenta':   pagadas.count(),
        'anio': hoy.year,
    }


@login_required
@require_GET
def platform_invoice_pdf(request, pk):
    """
    El PDF de una factura, para descargarlo o reenviarlo por otro camino.

    Lo alcanzan los dos niveles de plataforma: el soporte no emite ni cobra,
    pero atiende al cliente que dice no haber recibido nada, y para eso tiene
    que poder abrir el documento del que se habla.

    Se genera al vuelo y no se guarda. Una factura ya no cambia despues de
    emitida -- monto y plan quedan congelados -- asi que el PDF de hoy y el de
    dentro de un año son el mismo documento, salvo el sello de estado, que es
    precisamente lo que interesa que este al dia.
    """
    negado = _sin_permiso_de_plataforma(request.user)
    if negado:
        return negado

    factura = get_object_or_404(Invoice.objects.select_related('tenant'), pk=pk)
    respuesta = HttpResponse(generar_pdf_factura(factura),
                             content_type='application/pdf')
    respuesta['Content-Disposition'] = f'inline; filename="{factura.numero}.pdf"'
    return respuesta


@login_required
def platform_users(request):
    """
    Quién administra el SaaS. Solo el administrador de plataforma.

    Repartir este acceso es lo más crítico que hay en el panel: quien lo tiene
    puede dar de alta empresas y nombrar a sus administradores. Por eso no se
    delega en el soporte.
    """
    negado = _sin_permiso_de_plataforma(request.user, solo_admin=True)
    if negado:
        return negado

    msg, msg_is_error = '', False

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create':
            uname = request.POST.get('username', '').strip()
            pwd   = request.POST.get('password', '').strip()
            rol   = request.POST.get('role', 'staff')
            roles_validos = [r for r, _ in PLATFORM_ROLE_CHOICES]
            if not uname or not pwd:
                msg, msg_is_error = 'Username and password are required.', True
            elif rol not in roles_validos:
                msg, msg_is_error = f'Unknown platform role "{rol}".', True
            elif User.objects.filter(username=uname).exists():
                msg, msg_is_error = f'Username "{uname}" is already taken.', True
            else:
                # Sin tenant y sin UserProfile a propósito: un usuario de
                # plataforma no pertenece a ninguna empresa, y es esa ausencia la
                # que hace que get_tenant_or_404 lo deje fuera de las pantallas
                # del tenant. Tampoco es superusuario: el admin de Django y los
                # datos de las empresas siguen siendo otra cosa.
                with transaction.atomic():
                    u = User.objects.create_user(username=uname, password=pwd)
                    PlatformUser.objects.create(user=u, role=rol)
                msg = f'Platform user "{uname}" created as {rol}.'

        elif action == 'update_role':
            pk  = request.POST.get('platform_user_id')
            rol = request.POST.get('role', 'staff')
            acceso = get_object_or_404(PlatformUser, pk=pk)
            roles_validos = [r for r, _ in PLATFORM_ROLE_CHOICES]
            if rol not in roles_validos:
                msg, msg_is_error = f'Unknown platform role "{rol}".', True
            elif acceso.user == request.user:
                # Quitarse a uno mismo el nivel de administrador deja el panel
                # sin quien lo administre si es el único que queda.
                msg, msg_is_error = 'You cannot change your own platform role.', True
            else:
                acceso.role = rol
                acceso.save(update_fields=['role'])
                msg = f'"{acceso.user.username}" is now {rol}.'

        elif action == 'revoke':
            pk = request.POST.get('platform_user_id')
            acceso = get_object_or_404(PlatformUser, pk=pk)
            if acceso.user == request.user:
                msg, msg_is_error = 'You cannot revoke your own platform access.', True
            else:
                nombre = acceso.user.username
                # Se retira el acceso, no se borra la persona: el usuario sigue
                # existiendo y puede tener historial asociado.
                acceso.delete()
                msg = f'Platform access revoked for "{nombre}".'

    return render(request, 'warehouse/partials/platform_users.html', {
        'accesos': PlatformUser.objects.select_related('user').all(),
        'roles': PLATFORM_ROLE_CHOICES,
        'msg': msg, 'msg_is_error': msg_is_error,
        # Los superusuarios de Django no figuran en esta lista y conviene que se
        # vean: mientras no haya ningun administrador aqui, entran al panel; y
        # tengan o no acceso, siguen abriendo el admin de Django con los datos
        # de todas las empresas.
        'superusuarios': User.objects.filter(is_superuser=True).order_by('username'),
        'llave_maestra_activa': not PlatformUser.objects.filter(role='admin').exists(),
    })


@login_required
def platform_dashboard(request):
    """
    La pantalla del nivel de plataforma, aparte del tablero de las empresas.

    Tiene que ser una pagina propia y no una pestana mas: sus usuarios no
    pertenecen a ninguna empresa, asi que `dashboard` les responderia 404 al
    no encontrarles tenant. Que ese 404 exista es justo la garantia de que un
    usuario de plataforma no alcanza los datos de nadie.
    """
    negado = _sin_permiso_de_plataforma(request.user)
    if negado:
        return negado

    return render(request, 'warehouse/platform.html', {
        'platform_role': platform_role(request.user),
        'is_platform_admin': _es_admin_de_plataforma(request.user),
        # Quien ademas pertenece a una empresa puede volver a su tablero.
        'tiene_empresa': UserProfile.objects.filter(
            user=request.user, tenant__isnull=False).exists(),
    })


# ── EL HILO DE LA OPERACION ───────────────────────────────────────────────────

# Lo mas largo que se acepta en un mensaje. No es una regla de negocio, es un
# tope para que un pegado accidental no llene la pantalla ni la base.
MENSAJE_MAX = 4000


def lado_en_el_hilo(user):
    """
    En nombre de quien habla este usuario, o None si no le toca hablar.

    No es "quien eres" sino "de que lado escribes": del lado de la empresa
    contesta el operador del turno, sea staff o administrador, y del lado del
    cliente cualquiera de las personas que ese cliente tenga dadas de alta.
    Quien no tiene rol -el perfil vacio que devuelve get_profile- no escribe.
    """
    profile = get_profile(user)
    if profile.is_customer():
        return LADO_CLIENTE
    if profile.is_operator():
        return LADO_TENANT
    return None


def anotar_hilos(user, operaciones):
    """
    Le cuelga a cada operacion del listado si tiene hilo y cuantos mensajes
    sin leer le quedan a quien mira.

    Va en dos consultas y no en una por fila a proposito: el listado trae hasta
    doscientas operaciones, y preguntarle a cada una serian cuatrocientas
    consultas para pintar una tabla.

    Devuelve la lista ya anotada. El indicador es lo que hace que el hilo
    exista de verdad: sin el, un mensaje del cliente solo se descubre abriendo
    la operacion que a nadie se le ocurre abrir.
    """
    ops = list(operaciones)
    if not ops:
        return ops

    ids = [op.pk for op in ops]
    con_hilo = set(Conversation.objects.filter(operation_id__in=ids)
                   .values_list('operation_id', flat=True))

    # Hasta donde leyo *este* usuario cada hilo. Sin marca, todo lo del otro
    # lado cuenta como nuevo, que es lo que hace visible el primer mensaje.
    leido_hasta = ConversationRead.objects.filter(
        conversation=OuterRef('conversation'), user=user).values('last_read_at')[:1]
    pendientes = (Message.objects
                  .filter(conversation__operation_id__in=ids)
                  .exclude(author_id=user.pk)
                  .annotate(leido_hasta=Subquery(leido_hasta))
                  .filter(Q(leido_hasta__isnull=True) |
                          Q(created_at__gt=F('leido_hasta')))
                  .values('conversation__operation_id')
                  .annotate(n=Count('pk')))
    conteo = {fila['conversation__operation_id']: fila['n'] for fila in pendientes}

    for op in ops:
        op.tiene_hilo = op.pk in con_hilo
        op.sin_leer = conteo.get(op.pk, 0)
    return ops


def _operacion_del_hilo(request, pk):
    """
    La operacion cuyo hilo se pide, comprobando que quien lo pide pueda verla.

    Devuelve (operacion, respuesta_de_error). Es el mismo criterio que
    `operation_detail` -el tenant del request y, si mira un cliente, solo sus
    operaciones-, porque quien puede abrir el expediente puede leer lo que se
    conversa sobre el, y quien no, no.
    """
    tenant = get_tenant_or_404(request)
    op = get_object_or_404(WarehouseOperation, pk=pk, tenant=tenant)
    if not customer_can_access_op(request.user, op):
        return None, HttpResponse('Permission denied.', status=403)
    if lado_en_el_hilo(request.user) is None:
        return None, HttpResponse('Permission denied.', status=403)
    return op, None


def _pintar_hilo(request, op, error=''):
    """
    El hilo tal como se ve, y de paso lo da por leido.

    Marcar la lectura aqui y no en una llamada aparte es lo correcto: esta
    vista es la que refresca el panel cada pocos segundos mientras alguien lo
    tiene abierto, asi que pedirla *es* estar mirandolo. Cuando el panel se
    cierra deja de pedirse y lo que llegue despues vuelve a contar como nuevo.
    """
    hilo = getattr(op, 'conversation', None)
    mensajes = []
    if hilo:
        mensajes = list(hilo.messages.select_related('author'))
        hilo.marcar_leida(request.user)
    return render(request, 'warehouse/partials/chat_messages.html', {
        'operation': op,
        'mensajes': mensajes,
        'mi_lado': lado_en_el_hilo(request.user),
        'error': error,
    })


@login_required
def operation_chat(request, pk):
    """El hilo de una operacion. La pide el panel al abrirse y cada refresco."""
    op, negado = _operacion_del_hilo(request, pk)
    if negado:
        return negado
    return _pintar_hilo(request, op)


@login_required
@require_POST
def operation_chat_send(request, pk):
    """
    Escribe un mensaje en el hilo.

    El hilo se crea aqui, con el primer mensaje: una operacion sobre la que
    nadie ha dicho nada no necesita una fila esperando por si acaso.
    """
    op, negado = _operacion_del_hilo(request, pk)
    if negado:
        return negado

    cuerpo = (request.POST.get('body') or '').strip()
    if not cuerpo:
        return _pintar_hilo(request, op, error='Escribe un mensaje.')
    if len(cuerpo) > MENSAJE_MAX:
        return _pintar_hilo(request, op,
                            error=f'El mensaje no puede pasar de {MENSAJE_MAX} caracteres.')

    lado = lado_en_el_hilo(request.user)
    hilo, _ = Conversation.objects.get_or_create(
        operation=op, defaults={'tenant': op.tenant})

    ahora = timezone.now()
    mensaje = Message.objects.create(
        conversation=hilo,
        author=request.user,
        author_name=_nombre_visible(request.user),
        side=lado,
        body=cuerpo,
    )
    hilo.last_message_at = ahora
    hilo.save(update_fields=['last_message_at'])

    # Quien escribe ya leyo todo lo anterior, incluido lo que acaba de mandar.
    hilo.marcar_leida(request.user, ahora)

    notifications.avisar_mensaje_nuevo(hilo, lado, mensaje=mensaje,
                                       triggered_by=request.user)

    return _pintar_hilo(request, op)


def _nombre_visible(user):
    """
    Como se firma un mensaje. El nombre completo si lo hay, y si no el usuario.
    """
    completo = (user.get_full_name() or '').strip()
    return completo or user.username


@login_required
@require_GET
def chat_badges(request):
    """
    Cuantos mensajes sin leer tiene cada hilo, para refrescar los avisos de la
    tabla sin recargar la pantalla.

    La tabla se pinta una vez y se queda quieta: sin esto, un mensaje que llega
    no se ve hasta que alguien recarga, y el aviso de un hilo recien leido sigue
    encendido hasta la recarga siguiente. Recargar la tabla entera cada pocos
    segundos no es opcion -- son doscientas filas, un scroll horizontal a mano y
    varios menus abiertos -- asi que solo viajan los numeros.

    Devuelve unicamente las operaciones que quien pregunta puede ver y que ya
    tienen hilo; las demas no tienen nada que encender.
    """
    tenant = get_tenant_or_404(request)
    if lado_en_el_hilo(request.user) is None:
        return JsonResponse({'hilos': {}})

    ops = WarehouseOperation.objects.filter(
        tenant=tenant, conversation__isnull=False)
    ops = customer_ops_filter(request.user, ops)[:500]
    anotadas = anotar_hilos(request.user, ops)
    return JsonResponse({'hilos': {str(op.pk): op.sin_leer for op in anotadas}})
